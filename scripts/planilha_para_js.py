# -*- coding: utf-8 -*-
"""
Gera js/dados.js a partir de data/painel-processos-dados.xlsx — o fallback
embutido que o painel usa quando não consegue buscar a planilha (aberto
localmente via file://, sem servidor, ou se o fetch falhar). Rode depois de
editar a planilha real. Depois de scripts/atualizar_planilha.py, rode
scripts/cachear_formulas.py ANTES deste — atualizar_planilha.py salva com
openpyxl, que descarta o valor calculado das fórmulas já existentes (só
regrava a fórmula em si); sem recachear, colunas como Processos.Percentual e
Riscos.Nivel_PxI saem vazias abaixo:

    python scripts/atualizar_planilha.py
    python scripts/cachear_formulas.py
    python scripts/planilha_para_js.py

Lê as mesmas abas que js/app.js busca (CONFIG.abas) e escreve cada uma como
um array de objetos, no formato que o painel já espera.

Limitação conhecida: colunas com fórmula (ex. Riscos.Nivel_PxI, NUGEP
.Unidade_Nome) só saem com valor se algum programa com motor de planilha
(Excel, LibreOffice, Google Sheets) já tiver aberto e salvo o arquivo pelo
menos uma vez — o openpyxl não calcula fórmulas, só lê o último valor
armazenado nelas.
"""
import json
import os
import datetime as dt
from openpyxl import load_workbook

RAIZ = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(RAIZ, "..", "data", "painel-processos-dados.xlsx")
SAIDA = os.path.join(RAIZ, "..", "js", "dados.js")

# Mesma lista de abas de dados que js/app.js (CONFIG.abas) lê da planilha —
# se uma aba nova entrar lá, entre com ela aqui também.
ABAS = ['Macroprocessos', 'Processos', 'Subprocessos', 'Atividades', 'Tarefas',
        'Documentos', 'Riscos', 'Metricas', 'Medicoes', 'Papeis', 'Regras',
        'Cultura_Processos', 'Iniciativas', 'Competencias',
        'Jornada', 'Repositorio', 'NUGEP', 'Glossario', 'FAQ', 'Siglas', 'Parametros']


def valor_json(v):
    if isinstance(v, (dt.date, dt.datetime)):
        return v.isoformat()
    return v.strip() if isinstance(v, str) else v


def aba_para_lista(ws):
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []
    cabecalho = [str(h).strip() if h is not None else "" for h in linhas[0]]
    saida = []
    for linha in linhas[1:]:
        primeiro = linha[0] if linha else None
        if primeiro is None or str(primeiro).strip() == "":
            continue  # mesma regra do carregarXlsx() do painel: ignora linha sem valor na 1ª coluna
        obj = {}
        for h, v in zip(cabecalho, linha):
            if h:
                obj[h] = valor_json(v)
        saida.append(obj)
    return saida


def principal():
    wb = load_workbook(XLSX, data_only=True)
    dados = {}
    for aba in ABAS:
        dados[aba] = aba_para_lista(wb[aba]) if aba in wb.sheetnames else []
    dados["_gerado_em"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    corpo = json.dumps(dados, ensure_ascii=False, indent=1)
    with open(SAIDA, "w", encoding="utf-8") as f:
        f.write(
            "/* GERADO AUTOMATICAMENTE a partir de data/painel-processos-dados.xlsx — não edite à mão.\n"
            "   Fonte: data/painel-processos-dados.xlsx */\n"
            "window.PAINEL_DADOS = " + corpo + ";\n"
        )
    print(f"js/dados.js regenerado a partir de {len(ABAS)} abas de {XLSX}.")


if __name__ == "__main__":
    principal()
