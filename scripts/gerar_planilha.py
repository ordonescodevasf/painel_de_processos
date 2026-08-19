# -*- coding: utf-8 -*-
"""
Gera data/painel-processos-dados.xlsx com DADOS FICTÍCIOS — recria do zero
a planilha de exemplo.

ATENÇÃO — este script RECRIA a planilha do zero. Se o arquivo em data/ já
tem conteúdo real, NÃO rode este script: use scripts/atualizar_planilha.py.

Cabeçalhos, larguras de coluna e listas suspensas vêm de
scripts/esquema_planilha.py — a fonte única do esquema.

Códigos (PP-/SP-/AT-/TR-) reiniciam em 01 a cada novo vínculo com o pai —
não são mais únicos sozinhos (ver esquema_planilha.py). Trilha e as colunas
de rollup (Macroprocesso/Processo/Subprocesso/Atividade) são calculadas
aqui mesmo, na montagem das tuplas de dados, e gravadas como fórmula
autocontida (="texto") — não um VLOOKUP por código, que seria ambíguo.

Conteúdo de Jornada, Repositorio, NUGEP, Glossario, Competencias, FAQ e
Parametros vem de scripts/dados_conteudo.py.
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import esquema_planilha as ESQ
import dados_conteudo as CONTEUDO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AZUL = "0040B4"
AZUL_ESCURO = "0C326F"
CINZA_FORMULA = "F0F0F0"
CINZA_ZEBRA = "F5F8FC"
FONTE = "Arial"

COR_GUIA = {
    "LEIA-ME": AZUL_ESCURO,
    "Macroprocessos": AZUL, "Processos": AZUL, "Subprocessos": AZUL,
    "Atividades": AZUL, "Tarefas": AZUL,
    "Documentos": "168821", "Riscos": "168821", "Metricas": "168821",
    "Medicoes": "168821", "Papeis": "168821", "Regras": "168821",
    "Cultura_Processos": "168821", "Iniciativas": "168821",
    "Jornada": "155BCB", "Repositorio": "155BCB", "NUGEP": "155BCB",
    "Competencias": "155BCB", "Glossario": "155BCB", "FAQ": "155BCB",
    "Siglas": "888888", "Parametros": "888888", "Listas": "888888",
}

th = Side(style="thin", color="CCCCCC")
BORDA = Border(left=th, right=th, top=th, bottom=th)
F_HEAD = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
F_CELL = Font(name=FONTE, size=10)
FILL_HEAD = PatternFill("solid", fgColor=AZUL)
FILL_FORM = PatternFill("solid", fgColor=CINZA_FORMULA)
FILL_ZEBRA = PatternFill("solid", fgColor=CINZA_ZEBRA)
AL_HEAD = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_WRAP = Alignment(vertical="top", wrap_text=True)
AL_TOP = Alignment(vertical="top")
AL_CENTER = Alignment(horizontal="center", vertical="top")


def cabecalho(ws, aba, formula_cols=()):
    headers = ESQ.colunas(aba)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = AL_HEAD; c.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = ESQ.largura(aba, h)
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws._formula_cols = set(formula_cols)
    ws.sheet_properties.tabColor = COR_GUIA.get(aba, AZUL)
    ws.sheet_view.zoomScale = 90


def escreve(ws, linhas, wrap_cols=(), center_cols=()):
    for i, linha in enumerate(linhas, start=2):
        for j, v in enumerate(linha, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = F_CELL; c.border = BORDA
            if j in ws._formula_cols:
                c.fill = FILL_FORM
            elif i % 2 == 0:
                c.fill = FILL_ZEBRA
            c.alignment = AL_WRAP if j in wrap_cols else (AL_CENTER if j in center_cols else AL_TOP)


def dv(ws, col_letter, ref_, ultima_linha=300, strict=True):
    v = DataValidation(type="list", formula1=ref_, allow_blank=True,
        showErrorMessage=True, errorStyle="stop" if strict else "warning",
        errorTitle="Valor fora da lista",
        error="Escolha uma das opções da seta ao lado da célula.",
        showInputMessage=True, promptTitle="Lista de opções",
        prompt="Clique na seta ao lado da célula e escolha uma das opções.")
    ws.add_data_validation(v)
    v.add(f"{col_letter}2:{col_letter}{ultima_linha}")


def col(aba, nome):
    return get_column_letter(ESQ.colunas(aba).index(nome) + 1)


def lit_formula(ws, aba, campo, linhas):
    """Grava, em cada linha, uma fórmula autocontida (="valor já calculado")
    — usada em Trilha e nos rollups, onde um VLOOKUP por código seria
    ambíguo (o código reinicia a cada pai, não é mais único sozinho)."""
    j = ESQ.colunas(aba).index(campo) + 1
    for i, linha in enumerate(linhas, start=2):
        valor = linha.get(campo) if isinstance(linha, dict) else None
        if valor is None:
            continue
        texto = str(valor).replace('"', '""')
        c = ws.cell(row=i, column=j)
        c.value = f'="{texto}"'


def nivel_formula(cellref):
    c = cellref
    return (f'=IF(LEFT({c},2)="MG","Macroprocesso",IF(LEFT({c},2)="MF","Macroprocesso",'
            f'IF(LEFT({c},2)="MS","Macroprocesso",IF(LEFT({c},2)="PP","Processo",'
            f'IF(LEFT({c},2)="SP","Subprocesso",IF(LEFT({c},2)="AT","Atividade",'
            f'IF(LEFT({c},2)="TR","Tarefa","")))))))')


def aplicar_formula(ws, aba, campo, calc, number_format=None):
    j = ESQ.colunas(aba).index(campo) + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ""):
            continue
        c = ws.cell(row=r, column=j)
        c.value = calc(r)
        if number_format:
            c.number_format = number_format


# ─────────────────────────── Dados (fictícios) ──────────────────────────
# Tuplas na MESMA ORDEM das colunas de esquema_planilha.py (Trilha e
# rollups vêm calculados como dict auxiliar logo abaixo de cada bloco, não
# dentro da tupla, para poder usar lit_formula()).
SIGLAS = [
    ("10ª/AJ", "Assessoria Jurídica Regional"),
    ("10ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("10ª/EAG", "Escritório de Apoio Técnico de Araguaína"),
    ("10ª/GB", "Gabinete da Superintendência Regional"),
    ("10ª/GGR", "Gerência de Gestão Regional"),
    ("10ª/GTR", "Gerência Técnica Regional"),
    ("10ª/SL", "Secretaria Regional de Licitações"),
    ("10ª/SR", "10ª Superintendência Regional"),
    ("11ª/AJ", "Assessoria Jurídica Regional"),
    ("11ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("11ª/GB", "Gabinete da Superintendência Regional"),
    ("11ª/GGR", "Gerência de Gestão Regional"),
    ("11ª/GGR/UOF", "Unidade Regional de Orçamento e Finanças"),
    ("11ª/GGR/UOF/SCO", "Setor de Contabilidade"),
    ("11ª/GGR/URA", "Unidade Regional de Administração"),
    ("11ª/GRD", "Gerência Regional de Infraestrutura"),
    ("11ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("11ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("11ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("11ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("11ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("11ª/SL", "Secretaria Regional de Licitações"),
    ("11ª/SR", "11ª Superintendência Regional"),
    ("12ª/AJ", "Assessoria Jurídica Regional"),
    ("12ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("12ª/EMO", "Escritório de Apoio Técnico de Mossoró"),
    ("12ª/GB", "Gabinete da Superintendência Regional"),
    ("12ª/GGR", "Gerência de Gestão Regional"),
    ("12ª/GTR", "Gerência Técnica Regional"),
    ("12ª/SL", "Secretaria Regional de Licitações"),
    ("12ª/SR", "12ª Superintendência Regional"),
    ("13ª/AJ", "Assessoria Jurídica Regional"),
    ("13ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("13ª/EPT", "Escritório de Apoio Técnico de Patos"),
    ("13ª/GB", "Gabinete da Superintendência Regional"),
    ("13ª/GGR", "Gerência de Gestão Regional"),
    ("13ª/GTR", "Gerência Técnica Regional"),
    ("13ª/SL", "Secretaria Regional de Licitações"),
    ("13ª/SR", "13ª Superintendência Regional"),
    ("14ª/AJ", "Assessoria Jurídica Regional"),
    ("14ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("14ª/ETA", "Escritório de Apoio Técnico de Tauá"),
    ("14ª/GB", "Gabinete da Superintendência Regional"),
    ("14ª/GGR", "Gerência de Gestão Regional"),
    ("14ª/GTR", "Gerência Técnica Regional"),
    ("14ª/SL", "Secretaria Regional de Licitações"),
    ("14ª/SR", "14ª Superintendência Regional"),
    ("15ª/AJ", "Assessoria Jurídica Regional"),
    ("15ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("15ª/GB", "Gabinete da Superintendência Regional"),
    ("15ª/GGR", "Gerência de Gestão Regional"),
    ("15ª/GTR", "Gerência Técnica Regional"),
    ("15ª/SL", "Secretaria Regional de Licitações"),
    ("15ª/SR", "15ª Superintendência Regional"),
    ("16ª/AJ", "Assessoria Jurídica Regional"),
    ("16ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("16ª/EPI", "Escritório de Apoio Técnico de Piumhi"),
    ("16ª/GB", "Gabinete da Superintendência Regional"),
    ("16ª/GGR", "Gerência de Gestão Regional"),
    ("16ª/GTR", "Gerência Técnica Regional"),
    ("16ª/SL", "Secretaria Regional de Licitações"),
    ("16ª/SR", "16ª Superintendência Regional"),
    ("1ª/AJ", "Assessoria Jurídica Regional"),
    ("1ª/CIG", "Centro Integrado de Recursos Pesqueiros e Aquicultura de Gorutuba"),
    ("1ª/CIM", "Centro Integrado de Recursos Pesqueiros e Aquicultura de Três Marias"),
    ("1ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("1ª/EAL", "Escritório de Apoio Técnico de Almenara"),
    ("1ª/GB", "Gabinete da Superintendência Regional"),
    ("1ª/GB/URC", "Unidade Regional de Comunicação"),
    ("1ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("1ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("1ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("1ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("1ª/GRD", "Gerência Regional de Infraestrutura"),
    ("1ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("1ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("1ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("1ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("1ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("1ª/GRG/UFN/SCB", "Setor de Cobrança"),
    ("1ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("1ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("1ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("1ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("1ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("1ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("1ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("1ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("1ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("1ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("1ª/SL", "Secretaria Regional de Licitações"),
    ("1ª/SR", "1ª Superintendência Regional"),
    ("2ª/AJ", "Assessoria Jurídica Regional"),
    ("2ª/CIX", "Centro Integrado de Recursos Pesqueiros e Aquicultura de Xique-Xique"),
    ("2ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("2ª/EGU", "Escritório de Apoio Técnico de Guanambi"),
    ("2ª/EIR", "Escritório de Apoio Técnico de Irecê"),
    ("2ª/GB", "Gabinete da Superintendência Regional"),
    ("2ª/GB/URC", "Unidade Regional de Comunicação"),
    ("2ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("2ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("2ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("2ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("2ª/GRD", "Gerência Regional de Infraestrutura"),
    ("2ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("2ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("2ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("2ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("2ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("2ª/GRG/UFN/SCB", "Setor de Cobrança"),
    ("2ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("2ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("2ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("2ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("2ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("2ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("2ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("2ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("2ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("2ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("2ª/SL", "Secretaria Regional de Licitações"),
    ("2ª/SR", "2ª Superintendência Regional"),
    ("2ª/UBA", "Unidade Regional Descentralizada de Barreiras"),
    ("2ª/UVC", "Unidade Regional Descentralizada de Vitória da Conquista"),
    ("3ª/AJ", "Assessoria Jurídica Regional"),
    ("3ª/CIB", "Centro Integrado de Recursos Pesqueiros e Aquicultura de Bebedouro"),
    ("3ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("3ª/GB", "Gabinete da Superintendência Regional"),
    ("3ª/GB/URC", "Unidade Regional de Comunicação"),
    ("3ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("3ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("3ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("3ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("3ª/GRD", "Gerência Regional de Infraestrutura"),
    ("3ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("3ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("3ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("3ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("3ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("3ª/GRG/UFN/SCB", "Setor de Cobrança"),
    ("3ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("3ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("3ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("3ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("3ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("3ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("3ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("3ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("3ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("3ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("3ª/SL", "Secretaria Regional de Licitações"),
    ("3ª/SR", "3ª Superintendência Regional"),
    ("4ª/AJ", "Assessoria Jurídica Regional"),
    ("4ª/CIT", "Centro Integrado de Recursos Pesqueiros e Aquicultura de Betume"),
    ("4ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("4ª/EPR", "Escritório de Apoio Técnico de Propriá"),
    ("4ª/GB", "Gabinete da Superintendência Regional"),
    ("4ª/GB/URC", "Unidade Regional de Comunicação"),
    ("4ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("4ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("4ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("4ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("4ª/GRD", "Gerência Regional de Infraestrutura"),
    ("4ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("4ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("4ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("4ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("4ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("4ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("4ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("4ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("4ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("4ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("4ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("4ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("4ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("4ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("4ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("4ª/SL", "Secretaria Regional de Licitações"),
    ("4ª/SR", "4ª Superintendência Regional"),
    ("5ª/AJ", "Assessoria Jurídica Regional"),
    ("5ª/CII", "Centro Integrado de Recursos Pesqueiros e Aquicultura de Itiúba"),
    ("5ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("5ª/EPE", "Escritório de Apoio Técnico de Penedo"),
    ("5ª/GB", "Gabinete da Superintendência Regional"),
    ("5ª/GB/URC", "Unidade Regional de Comunicação"),
    ("5ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("5ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("5ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("5ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("5ª/GRD", "Gerência Regional de Infraestrutura"),
    ("5ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("5ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("5ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("5ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("5ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("5ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("5ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("5ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("5ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("5ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("5ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("5ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("5ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("5ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("5ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("5ª/SL", "Secretaria Regional de Licitações"),
    ("5ª/SR", "5ª Superintendência Regional"),
    ("6ª/AJ", "Assessoria Jurídica Regional"),
    ("6ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("6ª/ESA", "Escritório de Representação de Salvador"),
    ("6ª/GB", "Gabinete da Superintendência Regional"),
    ("6ª/GB/URC", "Unidade Regional de Comunicação"),
    ("6ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("6ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("6ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("6ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("6ª/GRD", "Gerência Regional de Infraestrutura"),
    ("6ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("6ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("6ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("6ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("6ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("6ª/GRG/UFN/SCB", "Setor de Cobrança"),
    ("6ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("6ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("6ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("6ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("6ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("6ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("6ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("6ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("6ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("6ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("6ª/SL", "Secretaria Regional de Licitações"),
    ("6ª/SR", "6ª Superintendência Regional"),
    ("6ª/UPF", "Unidade Regional Descentralizada de Paulo Afonso"),
    ("7ª/AJ", "Assessoria Jurídica Regional"),
    ("7ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("7ª/EOE", "Escritório de Apoio Técnico de Oeiras"),
    ("7ª/EPA", "Escritório de Apoio Técnico de Parnaíba"),
    ("7ª/ERN", "Escritório de Apoio Técnico de São Raimundo Nonato"),
    ("7ª/GB", "Gabinete da Superintendência Regional"),
    ("7ª/GB/URC", "Unidade Regional de Comunicação"),
    ("7ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("7ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("7ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("7ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("7ª/GRD", "Gerência Regional de Infraestrutura"),
    ("7ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("7ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("7ª/GRD/UPS", "Unidade Regional de Projetos Especiais"),
    ("7ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("7ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("7ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("7ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("7ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("7ª/GRI", "Gerência Regional de Irrigação e Operações"),
    ("7ª/GRI/UEI", "Unidade Regional de Gestão dos Empreendimentos de Irrigação"),
    ("7ª/GRI/URP", "Unidade Regional de Apoio à Produção"),
    ("7ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("7ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("7ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("7ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("7ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("7ª/SL", "Secretaria Regional de Licitações"),
    ("7ª/SR", "7ª Superintendência Regional"),
    ("8ª/AJ", "Assessoria Jurídica Regional"),
    ("8ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("8ª/EIM", "Escritório de Apoio Técnico de Imperatriz"),
    ("8ª/EPD", "Escritório de Apoio Técnico de Presidente Dutra"),
    ("8ª/GB", "Gabinete da Superintendência Regional"),
    ("8ª/GB/URC", "Unidade Regional de Comunicação"),
    ("8ª/GRA", "Gerência Regional de Administração e Tecnologia"),
    ("8ª/GRA/UGP", "Unidade Regional de Gestão de Pessoas"),
    ("8ª/GRA/USA", "Unidade Regional de Patrimônio, Logística e Serviços Auxiliares"),
    ("8ª/GRA/UTI", "Unidade Regional de Tecnologia da Informação"),
    ("8ª/GRD", "Gerência Regional de Infraestrutura"),
    ("8ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("8ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("8ª/GRD/UPS", "Unidade Regional de Projetos Especiais"),
    ("8ª/GRG", "Gerência Regional de Estratégia e Finanças"),
    ("8ª/GRG/UCB", "Unidade Regional de Contabilidade"),
    ("8ª/GRG/UFN", "Unidade Regional de Finanças e Cobrança"),
    ("8ª/GRG/UMC", "Unidade Regional de Monitoramento e de Controle de Contratos e Convênios"),
    ("8ª/GRG/URO", "Unidade Regional Orçamentária"),
    ("8ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("8ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("8ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("8ª/GRS", "Gerência Regional de Suporte Integrado"),
    ("8ª/GRS/ULF", "Unidade Regional de Licenciamento Ambiental e Administração Fundiária"),
    ("8ª/SL", "Secretaria Regional de Licitações"),
    ("8ª/SR", "8ª Superintendência Regional"),
    ("9ª/AJ", "Assessoria Jurídica Regional"),
    ("9ª/Cogex", "Comitê de Gestão Executiva da Superintendência Regional"),
    ("9ª/GB", "Gabinete da Superintendência Regional"),
    ("9ª/GGR", "Gerência de Gestão Regional"),
    ("9ª/GGR/UOF", "Unidade Regional de Orçamento e Finanças"),
    ("9ª/GGR/UOF/SCO", "Setor de Contabilidade"),
    ("9ª/GGR/URA", "Unidade Regional de Administração"),
    ("9ª/GRD", "Gerência Regional de Infraestrutura"),
    ("9ª/GRD/UEP", "Unidade Regional de Estudos e Projetos"),
    ("9ª/GRD/UIP", "Unidade Regional de Implantação e Acompanhamento de Projetos"),
    ("9ª/GRR", "Gerência Regional de Revitalização e Desenvolvimento Territorial"),
    ("9ª/GRR/UDT", "Unidade Regional de Desenvolvimento Territorial"),
    ("9ª/GRR/UES", "Unidade Regional de Empreendimentos Socioambientais"),
    ("9ª/SL", "Secretaria Regional de Licitações"),
    ("9ª/SR", "9ª Superintendência Regional"),
    ("AA", "Área de Administração e Tecnologia"),
    ("AA/GGP", "Gerência de Gestão de Pessoas"),
    ("AA/GGP/UBS", "Unidade de Benefícios e Saúde Ocupacional"),
    ("AA/GGP/UCP", "Unidade de Cadastro e Pagamento"),
    ("AA/GGP/UDP", "Unidade de Desenvolvimento de Pessoas"),
    ("AA/GGP/URT", "Unidade de Relações de Trabalho"),
    ("AA/GPA", "Gerência de Patrimônio, Logística e Serviços Auxiliares"),
    ("AA/GPA/SBI", "Setor de Biblioteca"),
    ("AA/GPA/SDP", "Setor de Documentação e Protocolo"),
    ("AA/GPA/UAL", "Unidade de Administração Predial e Logística"),
    ("AA/GPA/UCS", "Unidade de Compras Administrativas"),
    ("AA/GPA/UPM", "Unidade de Patrimônio e Material"),
    ("AA/GTI", "Gerência de Tecnologia da Informação"),
    ("AA/GTI/UIT", "Unidade de Infraestrutura e Tecnologia"),
    ("AA/GTI/UPC", "Unidade de Conformidade e Controle de Tecnologia da Informação"),
    ("AA/GTI/USC", "Unidade de Segurança Cibernética"),
    ("AA/GTI/USI", "Unidade de Sistemas de Informações"),
    ("AD", "Área de Desenvolvimento e Infraestrutura"),
    ("AD/GEP", "Gerência de Estudos e Projetos"),
    ("AD/GEP/UPE", "Unidade de Estudos e Projetos de Infraestrutura Urbana e Edificações"),
    ("AD/GEP/UPH", "Unidade de Estudos e Projetos de Infraestrutura Hídrica e Irrigação"),
    ("AD/GIM", "Gerência de Implantação de Obras"),
    ("AD/GIM/UOE", "Unidade de Implantação de Obras de Infraestrutura Urbana e Edificações"),
    ("AD/GIM/UOH", "Unidade de Implantação de Obras de Infraestrutura Hídrica e Irrigação"),
    ("AD/GPI", "Gerência de Parcerias e Investimentos"),
    ("AD/GQV", "Gerência de Qualificação Viária"),
    ("AD/GQV/UGV", "Unidade de Gestão e Controle de Projetos de Qualificação Viária"),
    ("AD/GQV/UIO", "Unidade de Implantação de Obras de Qualificação Viária"),
    ("AD/SE", "Secretaria Executiva"),
    ("AE", "Área de Estratégia e Finanças"),
    ("AE/GCB", "Gerência de Contabilidade"),
    ("AE/GCB/UCC", "Unidade de Conformidade Contábil"),
    ("AE/GCB/UEC", "Unidade de Execução Contábil"),
    ("AE/GCB/UER", "Unidade de Escrituração e Revisão Contábil"),
    ("AE/GFN", "Gerência de Finanças"),
    ("AE/GFN/UCR", "Unidade de Gestão da Cobrança"),
    ("AE/GFN/UEF", "Unidade de Programação e Execução Financeira"),
    ("AE/GGO", "Gerência de Gestão Orçamentária"),
    ("AE/GGO/UEO", "Unidade de Execução Orçamentária"),
    ("AE/GGO/UPO", "Unidade de Programação Orçamentária"),
    ("AE/GPE", "Gerência de Planejamento Estratégico"),
    ("AE/GPE/UAT", "Unidade de Acompanhamento de Transferências"),
    ("AE/GPE/UNP", "Unidade de Gestão Normativa e de Processos"),
    ("AE/GPE/UPL", "Unidade de Planejamento Institucional"),
    ("AG", "Área de Governança e Sustentabilidade"),
    ("AG/GAF", "Gerência de Administração Fundiária e Geotecnologia"),
    ("AG/GAF/UDF", "Unidade de Documentação Fundiária"),
    ("AG/GAF/UGG", "Unidade de Gestão Geotecnológica"),
    ("AG/GCT", "Gerência de Custos"),
    ("AG/GCT/UCT", "Unidade de Custos e Orçamentação"),
    ("AG/GCT/UTR", "Unidade de Procedimentos Técnicos Referenciais"),
    ("AG/GMA", "Gerência de Meio Ambiente"),
    ("AG/GMA/UCA", "Unidade de Conservação Ambiental"),
    ("AG/GMA/ULA", "Unidade de Licenciamento Ambiental"),
    ("AG/SE", "Secretaria Executiva"),
    ("AGE", "Assembleia Geral"),
    ("AI", "Área de Irrigação e Operações"),
    ("AI/GAP", "Gerência de Apoio à Produção Irrigada"),
    ("AI/GAP/UGI", "Unidade de Gestão de Áreas Irrigadas"),
    ("AI/GAP/UPI", "Unidade de Apoio aos Projetos Públicos de Irrigação"),
    ("AI/GEE", "Gerência de Eficiência Energética"),
    ("AI/GEE/UEE", "Unidade de Gestão de Energia e Apoio ao PISF"),
    ("AI/GEE/USB", "Unidade de Gestão e Segurança de Barragens"),
    ("AI/GGE", "Gerência de Gestão de Empreendimentos"),
    ("AI/GGE/UGM", "Unidade de Gestão e Monitoramento de Empreendimentos"),
    ("AI/GGE/UOR", "Unidade de Operação e Reabilitação de Empreendimentos"),
    ("AI/SE", "Secretaria Executiva"),
    ("AR", "Área de Revitalização e Desenvolvimento Territorial"),
    ("AR/GDT", "Gerência de Desenvolvimento Territorial"),
    ("AR/GDT/UAP", "Unidade de Desenvolvimento dos Arranjos Produtivos"),
    ("AR/GDT/UPA", "Unidade de Recursos Pesqueiros, Aquicultura e Beneficiamento"),
    ("AR/GMT", "Gerência de Mecanização e Modernização Territorial"),
    ("AR/GMT/UCM", "Unidade de Gestão e Controle da Modernização Territorial"),
    ("AR/GMT/UME", "Unidade de Aquisição de Máquinas e Equipamentos"),
    ("AR/GSA", "Gerência de Saneamento e Acesso à Água"),
    ("AR/GSA/UAS", "Unidade de Água e Saneamento Básico"),
    ("AR/GSA/UEA", "Unidade de Esgotamento Rural e Acesso a Água"),
    ("AR/SE", "Secretaria Executiva"),
    ("CEC", "Comissão de Ética da Codevasf"),
    ("Coaud", "Comitê de Auditoria Estatutário"),
    ("Cogid", "Comitê de Governança Interna e Digital"),
    ("Cogid/SPDP", "Subcomitê de Proteção de Dados Pessoais"),
    ("Cogid/SSI", "Subcomitê de Segurança da Informação"),
    ("Cogid/STIC", "Subcomitê de Tecnologia da Informação"),
    ("Confis", "Conselho Fiscal"),
    ("Consad", "Conselho de Administração"),
    ("Consad/AUD", "Auditoria Interna"),
    ("Consad/AUD/UGE", "Unidade de Acompanhamento e Avaliação da Gestão"),
    ("Consad/AUD/UIN", "Unidade de Apoio e Informação"),
    ("Consad/AUD/UPR", "Unidade de Acompanhamento e Avaliação de Programas"),
    ("Consad/COR", "Corregedoria"),
    ("Consad/COR/UIV", "Unidade Especial de Admissibilidade e Investigação"),
    ("Consad/COR/UPJ", "Unidade Especial de Processos Acusatórios e Julgamento"),
    ("Consad/OUV", "Ouvidoria"),
    ("Copes", "Comitê de Pessoas, Elegibilidade, Sucessão e Remuneração"),
    ("DEX", "Diretoria Executiva"),
    ("PR", "Presidência"),
    ("PR/AC", "Assessoria de Comunicação"),
    ("PR/AC/UCE", "Unidade Especial de Cerimonial e Eventos"),
    ("PR/AC/UIM", "Unidade Especial de Imagem e Promoção Institucional"),
    ("PR/AJ", "Assessoria Jurídica"),
    ("PR/AJ/UAA", "Unidade de Assuntos Administrativos"),
    ("PR/AJ/UCO", "Unidade do Contencioso"),
    ("PR/EBE", "Escritório de Representação de Belém"),
    ("PR/GB", "Gabinete da Presidência"),
    ("PR/SC", "Secretaria de Órgãos Colegiados"),
    ("PR/SI", "Secretaria de Integridade"),
    ("PR/SLC", "Secretaria de Licitações e Contratos"),
    ("PR/SLC/UGC", "Unidade Especial de Gestão de Contratos"),
    ("PR/SLC/UGL", "Unidade Especial de Gestão de Licitações"),
    ("PR/SRC", "Secretaria de Gestão de Riscos e Controle Interno"),
    ("PR/SRC/UGR", "Unidade Especial de Gerenciamento de Riscos Corporativos"),
]

MACROS = [
    ('MG-01', 'Gestão Estratégica e Governança', 'Gerencial', '1', 'AE/GPE', 'Formulação, desdobramento e monitoramento da estratégia corporativa, da governança e do desempenho institucional.', 'Garantir que a atuação da Companhia esteja alinhada à estratégia, com decisões baseadas em evidências.', 'Plano Estratégico; Plano de Ação Anual; Relatórios de desempenho', 'Diretoria Executiva; Conselho de Administração; Unidades internas', 'Ministério supervisor; Órgãos de controle; Sociedade', 'e-Codevasf; Painéis de BI', None, None),
    ('MG-02', 'Gestão de Riscos, Integridade e Controles', 'Gerencial', '2', 'AE/GAG', 'Identificação, avaliação e tratamento de riscos corporativos, integridade e controles internos (2ª linha).', 'Assegurar razoável segurança para o alcance dos objetivos institucionais.', 'Política de Gestão de Riscos; Matriz de riscos corporativa; Plano de integridade', 'Alta administração; Gestores de 1ª linha', 'Auditoria interna; CGU; TCU', 'e-Codevasf; Sistema de gestão de riscos', None, None),
    ('MF-01', 'Desenvolvimento Territorial e Estruturação Produtiva', 'Finalístico', '3', 'AR/GDT', 'Apoio a arranjos produtivos, inclusão socioeconômica e estruturação de cadeias produtivas nos vales.', 'Promover o desenvolvimento regional integrado e sustentável.', 'Projetos de estruturação produtiva; Convênios e instrumentos de repasse', 'Produtores rurais; Cooperativas; Municípios', 'Parlamentares; Entes federados; Entidades parceiras', 'TransfereGov; e-Codevasf', None, None),
    ('MF-02', 'Gestão de Empreendimentos de Irrigação', 'Finalístico', '4', 'AI/GOM', 'Implantação, operação, manutenção e transferência de gestão de perímetros públicos de irrigação.', 'Ampliar a área irrigada produtiva com sustentabilidade hídrica e econômica.', 'Perímetros em operação; Água distribuída; Relatórios de O&M', 'Irrigantes; Distritos de irrigação', 'ANA; Agências estaduais; Associações de usuários', 'SIG-Irrigação (fictício); e-Codevasf', None, None),
    ('MF-03', 'Revitalização de Bacias Hidrográficas', 'Finalístico', '5', 'AR/GRB', 'Ações de recuperação hidroambiental, segurança hídrica e uso sustentável dos recursos naturais.', 'Contribuir para a revitalização das bacias dos rios São Francisco e Parnaíba.', 'Nascentes recuperadas; Obras hidroambientais; Sistemas de abastecimento', 'Comunidades ribeirinhas; Municípios', 'Comitês de bacia; MMA; Órgãos ambientais', 'e-Codevasf; GeoPortal (fictício)', None, None),
    ('MS-01', 'Gestão de Licitações e Contratos', 'Suporte', '6', 'AA/GLC', 'Planejamento das contratações, seleção de fornecedores e gestão dos contratos administrativos da Companhia.', 'Prover contratações tempestivas, vantajosas e conformes à legislação.', 'Editais publicados; Contratos firmados; Atas de registro de preços', 'Todas as unidades demandantes', 'Fornecedores; Assessoria Jurídica; Órgãos de controle; PNCP', 'Compras.gov.br; PNCP; Painel de Preços; e-Codevasf', None, 'Macroprocesso priorizado no ciclo 2026 de mapeamento.'),
    ('MS-02', 'Gestão de Pessoas', 'Suporte', '7', 'AG/GGP', 'Provimento, desenvolvimento, remuneração e qualidade de vida do corpo funcional.', 'Assegurar pessoas qualificadas e engajadas para a missão institucional.', 'Empregados admitidos e capacitados; Folha de pagamento', 'Empregados; Gestores', 'Sindicatos; SEST; Ministério supervisor', 'SIGEP (fictício); e-Codevasf', None, None),
    ('MS-03', 'Gestão de Tecnologia da Informação', 'Suporte', '8', 'AT/GTI', 'Planejamento, desenvolvimento, sustentação e segurança dos serviços de TI.', 'Prover soluções digitais seguras que habilitem os processos de negócio.', 'Sistemas em produção; Serviços de infraestrutura; Suporte ao usuário', 'Todas as unidades', 'SGD/MGI; Fornecedores de TI', 'e-Codevasf; Service Desk', None, None),
]

# (Macroprocesso, Codigo, Nome, Descricao, Objetivo, Unidade_Organica_Responsavel,
#  Ponto_Focal_Nugep, Prioridade, Complexidade, Maturidade, Status_Mapeamento,
#  Inicio_Mapeamento, Prazo_Previsto, Data_Conclusao, Ultima_Atualizacao,
#  M1..M10, Fornecedores, Entradas, Saidas, Beneficiarios, Sistemas,
#  Processo_ECodevasf, Processo_ECodevasf_Link, Imagem_Bizagi,
#  Competencias_Necessarias, Fontes_Dados, Proxima_Acao, Pendencia)
PROCS_BASE = [
    ('MS-01', 'PP-01', 'Planejamento da Contratação', 'Da identificação da necessidade (DFD) até a aprovação do ETP, TR e pesquisa de preços que instruem o certame.', 'Instruir as contratações com estudos e artefatos completos, reduzindo retrabalho e impugnações.', 'AA/GLC', 'Carlos Eduardo Lima (UNP)', 'Alta', 'Alta', 'Definido', 'Concluído', '05/01/2026', '29/05/2026', '18/05/2026', '10/07/2026', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Áreas demandantes; Painel de Preços; Fornecedores (cotações)', 'DFD; Plano de Contratações Anual; Requisitos da área', 'ETP aprovado; TR aprovado; Pesquisa de preços validada', 'AA/GLC (Seleção do Fornecedor); Assessoria Jurídica', 'e-Codevasf; Compras.gov.br; Painel de Preços', '59500.000123/2026-11', 'https://e-processo.codevasf.gov.br/sei/controlador.php?acao=procedimento_visualizar&id=59500000123202611', 'https://fluxosti.codevasf.gov.br/incidentes/', 'Elaboração de estudos técnicos e termos de referência; Conhecimento da Lei nº 14.133/2021; Pesquisa e análise de preços de mercado', 'e-Codevasf; Painel de Preços; PNCP', 'Monitorar indicadores do processo e revisar o PRO em 12 meses.', None),
    ('MS-01', 'PP-02', 'Seleção do Fornecedor', 'Da divulgação do edital à homologação do resultado, incluindo sessão pública, julgamento e recursos.', 'Selecionar a proposta mais vantajosa com celeridade e segurança jurídica.', 'AA/GLC', 'Carlos Eduardo Lima (UNP)', 'Alta', 'Média', 'Repetível', 'Em andamento', '01/06/2026', '30/09/2026', None, '15/07/2026', 'Sim', 'Sim', 'Sim', 'Sim', 'Em andamento', 'Não', 'Não', 'Não', 'Não', 'Não', 'P-06.01 (artefatos); Compras.gov.br', 'Edital minutado; Parecer jurídico; ETP/TR', 'Resultado homologado; Contrato/ata para assinatura', 'Unidades demandantes; Fornecedores', 'Compras.gov.br; PNCP; e-Codevasf', '59500.000456/2026-22', 'https://e-processo.codevasf.gov.br/sei/controlador.php?acao=procedimento_visualizar&id=59500000456202622', None, 'Condução de sessões públicas de licitação; Julgamento de propostas e habilitação; Uso de plataformas de compras governamentais', 'Compras.gov.br', 'Validar AS-IS com o dono do processo (reunião marcada).', 'Pendente retorno da área sobre fluxo de recursos administrativos.'),
    ('MS-01', 'PP-03', 'Gestão e Fiscalização Contratual', 'Da assinatura do contrato ao encerramento, incluindo fiscalização, medições, pagamentos e sanções.', 'Garantir a entrega do objeto contratado no prazo, custo e qualidade pactuados.', 'AA/GLC', 'Bruna Souza (UNP)', 'Alta', 'Alta', 'Gerenciado', 'Em andamento', '15/06/2026', '30/11/2026', None, '20/07/2026', 'Sim', 'Sim', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'P-06.02 (contrato); Fornecedores contratados', 'Contrato assinado; Cronograma; Garantias', 'Objeto recebido; Pagamentos efetuados; Termo de encerramento', 'Unidades demandantes; Fornecedores', 'e-Codevasf; Compras.gov.br', '59500.000789/2026-33', 'https://e-processo.codevasf.gov.br/sei/controlador.php?acao=procedimento_visualizar&id=59500000789202633', None, 'Fiscalização técnica e administrativa de contratos; Gestão de medições e pagamentos; Aplicação de sanções administrativas', 'e-Codevasf', 'Concluir modelagem AS-IS das medições e pagamentos.', None),
    ('MF-02', 'PP-01', 'Operação e Manutenção de Perímetros Irrigados', 'Programação e distribuição de água, manutenção da infraestrutura de uso comum e relacionamento com irrigantes.', 'Assegurar a distribuição hídrica confiável e a conservação dos ativos.', 'AI/GOM', 'Eduardo Martins (UNP)', 'Média', 'Alta', 'Inicial', 'Em andamento', '04/05/2026', '30/10/2026', None, '12/07/2026', 'Sim', 'Sim', 'Sim', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'ANA (outorgas); Distritos de irrigação', 'Outorga de uso; Demanda de irrigação; Plano de cultivo', 'Água distribuída; Infraestrutura mantida; Relatórios de O&M', 'Irrigantes; Distritos', 'SIG-Irrigação (fictício); e-Codevasf', '59500.000901/2026-44', 'https://e-processo.codevasf.gov.br/sei/controlador.php?acao=procedimento_visualizar&id=59500000901202644', None, 'Operação de sistemas hidráulicos; Manutenção de infraestrutura de irrigação; Relacionamento com associações de irrigantes', 'SIG-Irrigação (fictício)', 'Agendar oficina de validação do AS-IS com o distrito.', None),
    ('MF-03', 'PP-01', 'Recuperação de Nascentes e Matas Ciliares', 'Seleção de áreas, execução de cercamento e plantio, e monitoramento da regeneração.', 'Recuperar áreas degradadas prioritárias das bacias.', 'AR/GRB', 'Eduardo Martins (UNP)', 'Média', 'Média', 'Inicial', 'Não iniciado', None, '31/03/2027', None, '01/07/2026', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Municípios; Comitês de bacia', 'Diagnóstico hidroambiental; Termos de cooperação', 'Nascentes recuperadas; Relatórios de monitoramento', 'Comunidades; Órgãos ambientais', 'e-Codevasf; GeoPortal (fictício)', None, None, None, 'Diagnóstico e recuperação de áreas degradadas; Noções de engenharia ambiental; Articulação com comunidades e órgãos ambientais', None, 'Enviar formulário de levantamento à área (previsto ago/2026).', None),
    ('MG-01', 'PP-01', 'Formulação e Monitoramento do Planejamento Estratégico', 'Construção do plano estratégico, desdobramento em planos de ação e monitoramento periódico de resultados.', 'Manter a estratégia viva, monitorada e comunicada.', 'AE/GPE', 'Gustavo Pereira (UNP)', 'Alta', 'Média', 'Otimizado', 'Concluído', '01/09/2025', '27/02/2026', '20/02/2026', '30/06/2026', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Sim', 'Diretoria; Unidades', 'Diretrizes de governo; Diagnóstico institucional', 'Plano Estratégico; Painel de indicadores; RAG', 'Diretoria Executiva; Conselhos', 'e-Codevasf; Painéis de BI', '59500.000015/2025-77', 'https://e-processo.codevasf.gov.br/sei/controlador.php?acao=procedimento_visualizar&id=59500000015202577', None, 'Planejamento estratégico e gestão por indicadores; Facilitação de oficinas; Análise de dados institucionais', None, 'Ciclo de monitoramento trimestral (próximo: set/2026).', None),
    ('MS-02', 'PP-01', 'Admissão e Integração de Empregados', 'Da homologação do concurso à integração do novo empregado, incluindo exames, posse e ambientação.', 'Admitir e integrar novos empregados com agilidade e conformidade.', 'AG/GGP', 'Bruna Souza (UNP)', 'Baixa', 'Baixa', 'Inicial', 'Não iniciado', None, '30/06/2027', None, '15/06/2026', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Não', 'Banca do concurso; Candidatos', 'Resultado homologado; Documentação do candidato', 'Empregado admitido e integrado', 'Novos empregados; Unidades de lotação', 'SIGEP (fictício); e-Codevasf', None, None, None, 'Rotinas de admissão e legislação trabalhista (CLT); Condução de processos de integração; Uso de sistemas de gestão de pessoas', None, 'Aguardando priorização no ciclo 2027.', None),
]

# (Macroprocesso, Processo, Codigo, Nome, Vinculo_Pai, Ordem, Descricao, Objetivo,
#  Unidade_Organica_Responsavel, Reutilizavel, Reutilizado_Em, Entradas, Saidas,
#  Sistemas, Fontes_Dados, Imagem_Bizagi)
SUBS_BASE = [
    ('MS-01', 'PP-01', 'SP-01', 'Estudo Técnico Preliminar (ETP)', 'PP-01', '1', 'Caracterização da necessidade, análise de soluções de mercado e demonstração da viabilidade da contratação.', 'Fundamentar tecnicamente a melhor solução para a necessidade.', 'AA/GLC', None, None, 'DFD; Levantamento de soluções de mercado', 'ETP aprovado no e-Codevasf', 'e-Codevasf; Compras.gov.br', 'e-Codevasf', None),
    ('MS-01', 'PP-01', 'SP-02', 'Termo de Referência (TR)', 'PP-01', '2', 'Definição do objeto, requisitos, modelo de execução e gestão contratual, e critérios de julgamento.', 'Especificar com precisão o objeto e as condições da contratação.', 'AA/GLC', None, None, 'ETP aprovado; Modelos padronizados de TR', 'TR aprovado e validado juridicamente', 'e-Codevasf', None, None),
    ('MS-01', 'PP-01', 'SP-03', 'Pesquisa de Preços', 'PP-01', '3', 'Levantamento de preços em fontes admitidas (Painel de Preços, PNCP, cotações) e consolidação do valor estimado.', 'Estimar o valor da contratação conforme a IN SEGES nº 65/2021.', 'AA/GLC', None, None, 'TR validado; Fontes de preços admitidas', 'Relatório de pesquisa de preços; Mapa comparativo', 'Painel de Preços; PNCP; e-Codevasf', 'Painel de Preços; PNCP', None),
    ('MS-01', 'PP-01', 'SP-01', 'Tratamento de Preços Inexequíveis ou Excessivos', 'SP-03', '1', 'Identificação e exclusão de valores discrepantes coletados na pesquisa, com justificativa técnica registrada.', 'Evitar que outliers de preço distorçam o valor estimado da contratação.', 'AA/GLC', 'Sim', 'MS-01/PP-02; MF-02/PP-01', 'Amostra de preços coletados na pesquisa', 'Valores discrepantes identificados e justificados', 'Planilha padrão; e-Codevasf', None, None),
    ('MS-01', 'PP-02', 'SP-01', 'Condução da Sessão Pública', 'PP-02', '1', 'Abertura da sessão no sistema, fase de lances, julgamento, habilitação e registro em ata.', 'Conduzir o certame com transparência e celeridade.', 'AA/GLC', None, None, 'Edital publicado; Propostas dos licitantes', 'Ata da sessão; Resultado por item', 'Compras.gov.br', None, None),
    ('MF-02', 'PP-01', 'SP-01', 'Distribuição de Água aos Irrigantes', 'PP-01', '1', 'Programação semanal, operação de comportas e bombas, e registro de volumes distribuídos.', 'Entregar a água programada com eficiência e equidade.', 'AI/GOM', None, None, 'Outorga de uso; Demanda de irrigação', 'Programação hídrica executada; Registros de volume', 'SIG-Irrigação (fictício)', None, None),
]

# (Macroprocesso, Processo, Subprocesso, Codigo, Nome, Vinculo_Pai, Ordem,
#  Tipo_Atividade, Descricao, Entradas, Saidas, Sistemas)
ATIVS_BASE = [
    ('MS-01', 'PP-01', 'SP-01', 'AT-01', 'Formalizar a necessidade (DFD)', 'SP-01', '1', 'Agregação de Valor', 'A área demandante registra a necessidade de contratação no Documento de Formalização da Demanda (DFD), justificando o objeto e vinculando-o ao Plano de Contratações Anual.', 'Necessidade identificada; Plano de Contratações Anual', 'DFD assinado no e-Codevasf', 'e-Codevasf'),
    ('MS-01', 'PP-01', 'SP-01', 'AT-02', 'Levantar soluções de mercado', 'SP-01', '2', None, 'A equipe de planejamento pesquisa catálogos, contratações similares no PNCP e alternativas de mercado que atendam à necessidade formalizada no DFD.', 'DFD; Catálogos; Contratações similares (PNCP)', 'Levantamento de soluções documentado', 'PNCP; Compras.gov.br'),
    ('MS-01', 'PP-01', 'SP-01', 'AT-03', 'Estimar quantidades e resultados', 'SP-01', '3', None, 'A equipe de planejamento estima quantitativos e resultados pretendidos com base no levantamento de soluções e em séries históricas de consumo.', 'Levantamento de soluções; Séries históricas', 'Memória de cálculo de quantitativos', 'e-Codevasf'),
    ('MS-01', 'PP-01', 'SP-01', 'AT-04', 'Elaborar e aprovar o ETP', 'SP-01', '4', 'Controle', 'A equipe de planejamento consolida os levantamentos no Estudo Técnico Preliminar (ETP) e o submete à aprovação da autoridade competente.', 'Levantamentos e memórias anteriores', 'ETP aprovado', 'e-Codevasf; Compras.gov.br'),
    ('MS-01', 'PP-01', 'SP-02', 'AT-01', 'Redigir o Termo de Referência', 'SP-02', '1', None, 'A equipe de planejamento redige a minuta do Termo de Referência a partir do ETP aprovado, utilizando os modelos padronizados da Empresa.', 'ETP aprovado; Modelos padronizados', 'Minuta de TR', 'e-Codevasf'),
    ('MS-01', 'PP-01', 'SP-02', 'AT-02', 'Validar o TR com a Assessoria Jurídica', 'SP-02', '2', 'Controle', 'A Assessoria Jurídica analisa a minuta do Termo de Referência e emite parecer sobre sua conformidade legal antes da publicação do edital.', 'Minuta de TR', 'TR validado; Parecer jurídico', 'e-Codevasf'),
    ('MS-01', 'PP-01', 'SP-03', 'AT-01', 'Realizar a pesquisa de preços', 'SP-03', '1', 'Agregação de Valor', 'A equipe de planejamento coleta preços em ao menos três fontes distintas (Painel de Preços, PNCP e contratações similares) para compor a estimativa de valor.', 'TR validado; Fontes de preços', 'Relatório de pesquisa de preços', 'Painel de Preços; PNCP'),
    ('MS-01', 'PP-01', 'SP-03', 'AT-02', 'Consolidar o valor estimado', 'SP-03', '2', None, 'A equipe de planejamento consolida os preços coletados e tratados estatisticamente em um valor estimado único para a contratação.', 'Relatório de pesquisa', 'Valor estimado consolidado', 'e-Codevasf'),
    ('MS-01', 'PP-01', 'SP-01', 'AT-01', 'Aplicar critério de exclusão de valores discrepantes', 'SP-01', '1', None, 'A equipe de planejamento aplica o critério estatístico definido para excluir da amostra valores inexequíveis ou excessivamente discrepantes, registrando a justificativa.', 'Amostra de preços coletados', 'Amostra tratada; Justificativa de exclusão', 'Planilha padrão'),
    ('MS-01', 'PP-02', 'SP-01', 'AT-01', 'Publicar o edital', 'SP-01', '1', 'Transferência', 'A área responsável publica o edital aprovado e o parecer jurídico no Compras.gov.br e no PNCP, abrindo o certame.', 'Edital aprovado; Parecer jurídico', 'Edital publicado (PNCP)', 'Compras.gov.br; PNCP'),
    ('MS-01', 'PP-02', 'SP-01', 'AT-02', 'Conduzir a sessão e julgar propostas', 'SP-01', '2', None, 'O agente de contratação conduz a sessão pública, analisa as propostas recebidas e registra o resultado do julgamento em ata.', 'Edital publicado; Propostas', 'Ata da sessão; Resultado do julgamento', 'Compras.gov.br'),
    ('MF-02', 'PP-01', 'SP-01', 'AT-01', 'Programar a distribuição hídrica', 'SP-01', '1', 'Agregação de Valor', 'A equipe de operação do perímetro programa semanalmente a distribuição de água com base no plano de cultivo e na disponibilidade hídrica.', 'Plano de cultivo; Disponibilidade hídrica', 'Programação semanal aprovada', 'SIG-Irrigação (fictício)'),
    ('MF-02', 'PP-01', 'SP-01', 'AT-02', 'Operar e registrar a distribuição', 'SP-01', '2', 'Agregação de Valor', 'A equipe de operação executa a distribuição programada e registra volumes e ocorrências no sistema de gestão da irrigação.', 'Programação semanal', 'Volumes registrados; Ocorrências', 'SIG-Irrigação (fictício)'),
    ('MS-01', 'PP-03', None, 'AT-01', 'Designar fiscais e gestor do contrato', 'PP-03', '1', 'Controle', 'Formalizar a designação do fiscal técnico, do fiscal administrativo e do gestor do contrato, com ciência dos designados.', 'Contrato assinado; Minuta de portaria de designação', 'Portaria de designação publicada', 'e-Codevasf'),
    ('MS-01', 'PP-03', None, 'AT-02', 'Registrar medição e atestar a execução', 'PP-03', '2', 'Agregação de Valor', 'Conferir o objeto entregue no período, registrar a medição e atestar a nota fiscal para pagamento.', 'Relatório de execução; Nota fiscal', 'Medição registrada; Nota fiscal atestada', 'e-Codevasf'),
]

# (Macroprocesso, Processo, Subprocesso, Atividade, Codigo, Nome, Ordem,
#  Tipo_Tarefa, Duracao_Estimada, Observacoes)
TAREFAS_BASE = [
    ('MS-01', 'PP-01', 'SP-01', 'AT-01', 'TR-01', 'Reunir informações da demanda', '1', 'Manual', 4, None),
    ('MS-01', 'PP-01', 'SP-01', 'AT-01', 'TR-02', 'Preencher o formulário DFD no e-Codevasf', '2', 'Manual', 4, 'Modelo DOC-006/DOC-012.'),
    ('MS-01', 'PP-01', 'SP-01', 'AT-01', 'TR-03', 'Colher assinatura eletrônica do gestor', '3', 'Manual', 8, None),
    ('MS-01', 'PP-01', 'SP-03', 'AT-01', 'TR-01', 'Consultar o Painel de Preços', '1', 'Manual', 4, None),
    ('MS-01', 'PP-01', 'SP-03', 'AT-01', 'TR-02', 'Consultar contratações no PNCP', '2', 'Manual', 4, None),
    ('MS-01', 'PP-01', 'SP-03', 'AT-01', 'TR-03', 'Registrar cotações de fornecedores', '3', 'Manual', 24, 'Mínimo de 3 fontes (IN 65/2021).'),
    ('MS-01', 'PP-01', 'SP-03', 'AT-01', 'TR-04', 'Aplicar tratamento estatístico', '4', 'Regra de negócio', 4, None),
    ('MS-01', 'PP-02', 'SP-01', 'AT-01', 'TR-01', 'Cadastrar o edital no Compras.gov.br', '1', 'Manual', 4, None),
    ('MS-01', 'PP-02', 'SP-01', 'AT-01', 'TR-02', 'Publicar o aviso no PNCP', '2', 'Automatizada', 0, None),
    ('MF-02', 'PP-01', 'SP-01', 'AT-01', 'TR-01', 'Consolidar demandas semanais dos lotes', '1', 'Manual', 8, None),
    ('MS-01', 'PP-03', None, 'AT-02', 'TR-01', 'Conferir o objeto entregue', '1', 'Manual', 8, None),
    ('MS-01', 'PP-03', None, 'AT-02', 'TR-02', 'Registrar a medição no processo', '2', 'Manual', 4, None),
    ('MS-01', 'PP-03', None, 'AT-02', 'TR-03', 'Atestar a nota fiscal', '3', 'Manual', 4, 'O prazo de pagamento conta a partir do ateste.'),
]

# (ID, Titulo, Tipo_Documento, Vinculo_Codigo, Versao, Data, Situacao, Ato_Aprovacao, Link, Observacoes)
DOCS_BASE = [
    ('DOC-001', 'PRO 06.01 — Planejamento da Contratação', 'Procedimento (PRO)', 'PP-01', '2.0', '18/05/2026', 'Vigente', 'Resolução nº 812, de 18/05/2026 (Diretoria Executiva)', 'https://exemplo.codevasf.gov.br/repositorio/pop-06-01.pdf', 'Publicado após validação do TO-BE.'),
    ('DOC-002', 'Diagrama AS-IS — Planejamento da Contratação (Bizagi)', 'Diagrama BPMN', 'PP-01', '1.0', '02/03/2026', 'Vigente', None, 'https://fluxosti.codevasf.gov.br/incidentes/', 'Exportado do Bizagi Modeler. Mesmo link da ficha do processo (P-06.01).'),
    ('DOC-003', 'Diagrama TO-BE — Planejamento da Contratação (Bizagi)', 'Diagrama BPMN', 'PP-01', '1.0', '04/05/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/repositorio/p-06-01-tobe.pdf', None),
    ('DOC-004', 'Ata — Reunião de contextualização com a GLC', 'Ata de reunião', 'PP-01', '1.0', '12/01/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/sei/ata-contextualizacao-p0601', 'e-Codevasf 59500.000123/2026-11.'),
    ('DOC-005', 'Relatório de oportunidades de melhoria (AS-IS → TO-BE)', 'Relatório', 'PP-01; SP-03', '1.0', '13/04/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/repositorio/rel-melhorias-p0601.pdf', None),
    ('DOC-006', 'Modelo de DFD — Documento de Formalização da Demanda', 'Formulário/Modelo', 'SP-01', '3.1', '10/02/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/modelos/dfd.docx', None),
    ('DOC-007', 'Roteiro de pesquisa de preços (IN SEGES nº 65/2021)', 'Procedimento (PRO)', 'SP-03; SP-01', '1.2', '27/04/2026', 'Vigente', 'Resolução nº 845, de 27/04/2026 (Diretor da Área de Aquisições)', 'https://exemplo.codevasf.gov.br/repositorio/roteiro-precos.pdf', None),
    ('DOC-008', 'NI-027/2024 — Norma de Licitações e Contratos (fictícia)', 'Norma interna', 'MS-01', '4.0', '20/11/2024', 'Vigente', 'Resolução nº 621, de 20/11/2024 (Diretoria Executiva)', 'https://exemplo.codevasf.gov.br/normativos/ni-027', None),
    ('DOC-009', 'Manual de Gestão e Fiscalização de Contratos (fictício)', 'Manual', 'PP-01; PP-03', '2.3', '14/08/2025', 'Em revisão', 'Resolução nº 703, de 14/08/2025 (Diretoria Executiva)', 'https://exemplo.codevasf.gov.br/normativos/manual-fiscalizacao', None),
    ('DOC-010', 'Ata — Contextualização do processo Seleção do Fornecedor', 'Ata de reunião', 'PP-02', '1.0', '02/06/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/sei/ata-contextualizacao-p0602', None),
    ('DOC-011', 'Diagrama AS-IS parcial — O&M de Perímetros (Bizagi)', 'Diagrama BPMN', 'PP-01', '0.3', '08/07/2026', 'Em elaboração', None, None, 'Modelagem em curso.'),
    ('DOC-012', 'Passo a passo do DFD no e-Codevasf', 'Formulário/Modelo', 'AT-01', '1.0', '12/02/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/modelos/dfd-sei-passoapasso.pdf', None),
    ('DOC-013', 'Caderno de Indicadores Estratégicos 2026', 'Relatório', 'PP-01', '1.0', '31/03/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/estrategia/caderno-2026.pdf', None),
    ('DOC-014', 'Plano de implantação do TO-BE (PMBOK — plano do projeto)', 'Plano', 'PP-01', '1.0', '06/05/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/repositorio/plano-tobe-p0601.pdf', None),
    ('DOC-015', 'Checklist de Conformidade — Planejamento da Contratação', 'Checklist', 'PP-01', '1.0', '18/05/2026', 'Vigente', None, 'https://exemplo.codevasf.gov.br/repositorio/checklist-p0601.pdf', 'Exemplo — checklist ainda não padronizado para os demais processos da carteira.'),
    ('DOC-016', 'NI-014/2019 — Norma de Administração de Pessoal (fictícia)', 'Norma interna', 'MS-02', '3.0', '10/06/2019', 'Vigente', 'Resolução nº 512, de 10/06/2019 (Diretoria Executiva)', 'https://exemplo.codevasf.gov.br/normativos/ni-014', 'Exemplo — norma com revisão trienal vencida (N-000, item 4.2.3), ainda não atualizada.'),
    ('DOC-017', 'Lei nº 13.303/2016 — Estatuto Jurídico das Estatais', 'Legislação externa', 'MG-01', '—', '30/06/2016', 'Vigente', None, 'http://www.planalto.gov.br/ccivil_03/_ato2015-2018/2016/lei/l13303.htm', 'Migrado da coluna Normativos_Aplicaveis do macroprocesso.'),
    ('DOC-018', 'Lei nº 14.133/2021 — Nova Lei de Licitações e Contratos', 'Legislação externa', 'MS-01', '—', '01/04/2021', 'Vigente', None, 'http://www.planalto.gov.br/ccivil_03/_ato2019-2022/2021/lei/L14133.htm', 'Migrado da coluna Normativos_Aplicaveis/Normativos_Relacionados.'),
    ('DOC-019', 'Lei nº 12.787/2013 — Política Nacional de Irrigação', 'Legislação externa', 'MF-02', '—', '11/01/2013', 'Vigente', None, 'http://www.planalto.gov.br/ccivil_03/_ato2011-2014/2013/lei/l12787.htm', 'Migrado da coluna Normativos_Aplicaveis do macroprocesso.'),
    ('DOC-020', 'CLT — Consolidação das Leis do Trabalho', 'Legislação externa', 'MS-02', '—', '01/05/1943', 'Vigente', None, 'http://www.planalto.gov.br/ccivil_03/decreto-lei/del5452.htm', 'Migrado da coluna Normativos_Aplicaveis do macroprocesso.'),
    ('DOC-021', 'Estatuto Social da Codevasf', 'Norma interna', 'MG-01', '—', None, 'Vigente', None, 'https://exemplo.codevasf.gov.br/normativos/estatuto-social', 'Migrado da coluna Normativos_Aplicaveis do macroprocesso.'),
    ('DOC-022', 'Regimento Interno da Codevasf', 'Norma interna', 'MG-01', '—', None, 'Vigente', None, 'https://exemplo.codevasf.gov.br/normativos/regimento-interno', 'Migrado da coluna Normativos_Aplicaveis do macroprocesso.'),
]

# (ID, Vinculo_Codigo, Descricao_Risco, Categoria, Fatores, Cronograma_Risco,
#  Data_Identificacao, Probabilidade_1a5, Impacto_1a5, Resposta,
#  Controles_Tratamento, Prazo_Tratamento, Responsavel, Status)
RISCOS_BASE = [
    ('RIS-001', 'PP-01', 'Estimativas de preço defasadas gerando sobrepreço ou licitação deserta.', 'Operacional', None, None, '05/01/2026', '4', '4', 'Mitigar', 'Roteiro de pesquisa com múltiplas fontes (Painel de Preços, PNCP) e revisão pela UNP.', '30/09/2026', 'Ricardo Nogueira', 'Em tratamento'),
    ('RIS-002', 'PP-01', 'TR genérico ou restritivo provocando impugnações e atrasos no certame.', 'Legal/Conformidade', None, None, '05/01/2026', '3', '4', 'Mitigar', 'Checklist de revisão e validação obrigatória pela Assessoria Jurídica.', '30/09/2026', 'Ricardo Nogueira', 'Em tratamento'),
    ('RIS-003', 'PP-03', 'Fiscalização intempestiva de contratos, com medições e pagamentos atrasados.', 'Operacional', None, None, '15/06/2026', '3', '5', 'Mitigar', 'Designação formal de fiscais, agenda de medições e alertas no e-Codevasf.', '30/11/2026', 'Ricardo Nogueira', 'Aberto'),
    ('RIS-004', 'SP-01', 'Conhecimento concentrado em um único empregado (pessoa-chave) na elaboração de ETP.', 'Pessoas', None, None, '10/02/2026', '4', '3', 'Mitigar', 'Publicar o PRO, treinar substitutos e revezar a equipe de planejamento.', '31/10/2026', 'Patrícia Ramos', 'Em tratamento'),
    ('RIS-005', 'PP-01', 'Indisponibilidade hídrica comprometendo a programação de distribuição.', 'Operacional', None, None, '04/05/2026', '2', '5', 'Mitigar', 'Plano de contingência hídrica e priorização de culturas conforme regras do perímetro.', None, 'Marcos Vinícius', 'Aberto'),
    ('RIS-006', 'PP-02', 'Instabilidade do Compras.gov.br durante a sessão pública.', 'Tecnologia da Informação', None, None, '01/06/2026', '2', '4', 'Mitigar', 'Protocolo de suspensão formal da sessão e comunicação imediata aos licitantes.', None, 'Daniela Ribeiro', 'Aberto'),
    ('RIS-007', 'MS-01', 'Alterações normativas frequentes exigindo atualização contínua de procedimentos.', 'Legal/Conformidade', None, None, '12/01/2026', '4', '2', 'Mitigar', 'Monitoramento normativo mensal (resenha) e revisão programada dos PROs.', '31/12/2026', 'Bruna Souza', 'Em tratamento'),
    ('RIS-008', 'AT-01', 'Uso de fontes de preço não admitidas pela IN SEGES nº 65/2021.', 'Legal/Conformidade', None, None, '10/01/2026', '2', '3', 'Mitigar', 'Validação da pesquisa pela UNP antes da aprovação do valor estimado.', '27/04/2026', 'Carlos Eduardo Lima', 'Encerrado'),
    ('RIS-009', 'PP-01', 'Baixo engajamento das áreas demandantes no preenchimento do DFD.', 'Pessoas', None, None, '05/01/2026', '3', '3', 'Mitigar', 'Oficinas de capacitação e modelo simplificado de DFD no e-Codevasf.', '31/08/2026', 'Bruna Souza', 'Encerrado'),
]

# (ID, Nome, Vinculo_Codigo, Categoria, Descricao_Formula, Criterios_Desempenho,
#  Unidade, Polaridade, Meta, Periodicidade, Fonte, Observacoes)
METRICAS_BASE = [
    ('IND-001', 'Tempo médio do ciclo de contratação', 'MS-01', 'Processo', 'Média de dias entre o DFD e a homologação do certame.', None, 'dias', 'Menor melhor', '120', 'Trimestral', 'Compras.gov.br', None),
    ('IND-002', 'Prazo médio de elaboração do ETP', 'PP-01', 'Processo', 'Média de dias úteis entre o DFD e a aprovação do ETP.', None, 'dias', 'Menor melhor', '30', 'Mensal', 'e-Codevasf', None),
    ('IND-003', 'Impugnações por edital', 'PP-02', 'Processo', 'Nº médio de impugnações recebidas por edital publicado.', None, 'nº', 'Menor melhor', '1', 'Trimestral', 'Compras.gov.br', None),
    ('IND-004', 'Contratos com fiscal designado', 'PP-03', 'Processo', '% de contratos vigentes com fiscal formalmente designado.', None, '%', 'Maior melhor', '100', 'Mensal', 'e-Codevasf', None),
    ('IND-005', 'Eficiência no uso da água', 'PP-01', 'Processo', 'Relação % entre volume faturado e volume captado no perímetro.', None, '%', 'Maior melhor', '75', 'Mensal', 'SIG-Irrigação (fictício)', None),
    ('IND-006', 'Processos com mapeamento concluído', 'MS-01', 'Processo', '% de processos do macroprocesso com marco M9 concluído.', None, '%', 'Maior melhor', '100', 'Trimestral', 'Painel de Processos', None),
    ('IND-007', 'Ações estratégicas monitoradas no prazo', 'PP-01', 'Processo', '% de ações do plano com status atualizado no ciclo.', None, '%', 'Maior melhor', '95', 'Trimestral', 'Painéis de BI', None),
    ('IND-008', 'Pesquisas de preço com 3+ fontes', 'SP-03', 'Processo', '% de pesquisas de preços com três ou mais fontes admitidas.', None, '%', 'Maior melhor', '100', 'Mensal', 'e-Codevasf', None),
    ('IND-009', 'Nascentes recuperadas no ano', 'MF-03', 'Processo', 'Nº de nascentes com recuperação concluída no exercício.', None, 'nº', 'Maior melhor', '120', 'Anual', 'GeoPortal (fictício)', None),
    ('MET-010', 'SLA — Tempo de resposta ao demandante', 'PP-01', 'Nível de Serviço (SLA)', 'Dias úteis entre o recebimento do DFD e o retorno inicial à área demandante.', None, 'dias úteis', 'Menor melhor', 5, 'Mensal', 'e-Codevasf', 'Acordado com as áreas demandantes em 2026.'),
    ('MET-011', 'ROI — Automatização da pesquisa de preços', 'PP-01', 'Financeiro (ROI)', '(Economia de horas de trabalho convertida em custo − custo da ferramenta) / custo da ferramenta.', None, '%', 'Maior melhor', 150, 'Anual', 'AE/GPE', 'Estimativa fictícia para demonstração.'),
]

MEDICOES = [
    ('MED-001', 'IND-001', '30/06/2026', '148', None),
    ('MED-002', 'IND-002', '30/06/2026', '26', None),
    ('MED-003', 'IND-003', '30/06/2026', '0.8', None),
    ('MED-004', 'IND-004', '30/06/2026', '92', None),
    ('MED-005', 'IND-005', '30/06/2026', '71', None),
    ('MED-006', 'IND-006', '30/06/2026', '33', None),
    ('MED-007', 'IND-007', '30/06/2026', '97', None),
    ('MED-008', 'IND-008', '30/06/2026', '88', None),
    ('MED-901', 'IND-002', '31/03/2026', 165, 'Primeira apuração do trimestre.'),
    ('MED-902', 'IND-002', '31/12/2025', 178, None),
    ('MED-903', 'MET-010', '31/07/2026', 4, None),
    ('MED-904', 'MET-010', '30/06/2026', 6, 'Acúmulo de demandas no fim do semestre.'),
    ('MED-905', 'MET-011', '30/06/2026', 132, None),
]

# (ID, Vinculo_Codigo, Papel, Envolvimento, Unidade_Pessoa)
PAPEIS_BASE = [
    ('PAP-001', 'PP-01', 'Gestor do processo', 'Aprova (A)', 'AE/GAG'),
    ('PAP-002', 'PP-01', 'Equipe de planejamento da contratação', 'Executa (R)', 'AR/GDT'),
    ('PAP-003', 'PP-01', 'Assessoria Jurídica', 'Consultado (C)', 'AA/GLC'),
    ('PAP-004', 'PP-01', 'UNP', 'Informado (I)', 'AE/GPE/UNP'),
    ('PAP-005', 'SP-03', 'Equipe de planejamento da contratação', 'Executa (R)', 'AR/GDT'),
    ('PAP-006', 'SP-03', 'UNP', 'Consultado (C)', 'AE/GPE/UNP'),
    ('PAP-007', 'AT-01', 'Analista de planejamento da contratação', 'Executa (R)', 'AR/GDT'),
    ('PAP-008', 'PP-02', 'Gestor do processo', 'Aprova (A)', 'Gerente de Licitações'),
    ('PAP-009', 'PP-03', 'Gestor do processo', 'Aprova (A)', 'Gerente de Licitações'),
    ('PAP-010', 'PP-01', 'Gestor do processo', 'Aprova (A)', 'Gerente de Operação'),
    ('PAP-011', 'PP-01', 'Gestor do processo', 'Aprova (A)', 'Gerente de Revitalização'),
    ('PAP-012', 'PP-01', 'Gestor do processo', 'Aprova (A)', 'Gerente de Planejamento'),
    ('PAP-013', 'PP-01', 'Gestor do processo', 'Aprova (A)', 'Gerente de Pessoas'),
]

# (ID, Nome, Vinculo_Codigo, Tipo_Regra, Descricao, Fonte_Normativa)
REGRAS_BASE = [
    ('REG-001', 'Alçada de aprovação por valor', 'MS-01', 'Operação', 'Contratações acima de R$ 50.000 exigem aprovação da autoridade competente.', 'Lei nº 14.133/2021'),
    ('REG-002', 'Pesquisa de preços obrigatória', 'PP-01', 'Regulatória', 'Toda contratação exige pesquisa de preços prévia, conforme métodos definidos em norma.', 'IN SEGES nº 65/2021'),
    ('REG-003', 'Exclusão de valores discrepantes', 'SP-03', 'Decisão', 'Valores inexequíveis ou excessivos devem ser excluídos da amostra, com justificativa registrada.', 'IN SEGES nº 65/2021, art. 6º'),
    ('REG-004', 'Validação jurídica obrigatória', 'AT-02', 'Procedimento e política', 'Toda minuta de Termo de Referência deve ser submetida à Assessoria Jurídica antes da publicação do edital.', 'Lei nº 14.133/2021, art. 53'),
]

CULTURA = [
    ('CULT-01', 1, 'Acordo geral sobre o que são os processos de negócio', 'Não avaliado', None),
    ('CULT-02', 2, 'Compreensão de como os processos de negócio interagem e se afetam uns aos outros', 'Não avaliado', None),
    ('CULT-03', 3, 'Definição clara do valor que cada processo produz', 'Não avaliado', None),
    ('CULT-04', 4, 'Documentação de como cada processo produz os seus resultados', 'Não avaliado', None),
    ('CULT-05', 5, 'Compreensão das competências necessárias para cada processo', 'Não avaliado', None),
    ('CULT-06', 6, 'Compreensão do desempenho de cada processo', 'Não avaliado', None),
    ('CULT-07', 7, 'Medição contínua do desempenho dos processos', 'Não avaliado', None),
    ('CULT-08', 8, 'Decisões de gestão baseadas no conhecimento do desempenho dos processos', 'Não avaliado', None),
    ('CULT-09', 9, 'Donos de cada processo com a responsabilidade pelo desempenho do processo', 'Não avaliado', None),
]

INICIATIVAS = [
    ('INI-001', '20/05/2026', 'Automação da conferência de preços na pesquisa (PP-01)', 'Automação', 'A equipe de TI usou o PRO e o roteiro de pesquisa de preços publicados no repositório (DOC-007) para especificar uma rotina que já sinaliza valores fora da faixa de mercado antes da aprovação do ETP.', 'PP-01'),
    ('INI-002', '10/06/2026', 'Nova gerência reaproveitou o modelo de DFD para abrir seu próprio mapeamento', 'Reaproveitamento de modelo', 'A Gerência de Pessoas usou o Modelo de DFD (DOC-006) e o roteiro de contextualização do repositório para estruturar o levantamento do processo de Admissão e Integração (PP-07), sem precisar da UNP para montar os artefatos do zero.', 'PP-01'),
    ('INI-003', '02/07/2026', 'Checklist de conformidade virou exigência do comitê de contratações', 'Decisão de governança', 'O Checklist de Conformidade (DOC-015), criado para o PP-01, foi adotado pelo comitê interno de contratações como item obrigatório de pauta antes da homologação de qualquer certame acima de R$ 100 mil.', 'PP-01; PP-02'),
]


def _nivel(codigo):
    c = str(codigo or "").split(";")[0].strip()
    if c[:2] in ("MG", "MF", "MS"):
        return "Macroprocesso"
    if c.startswith("PP"):
        return "Processo"
    if c.startswith("SP"):
        return "Subprocesso"
    if c.startswith("AT"):
        return "Atividade"
    if c.startswith("TR"):
        return "Tarefa"
    return ""


# Monta os dicts completos (Trilha/rollups incluídos) a partir das tuplas BASE.
PROCS = []
for (mp, cod, nome, desc, obj, uor, pfn, pri, com, mat, sm, im, pp, dc_, ua,
     m1, m2, m3, m4, m5, m6, m7, m8, m9, m10, forn, ent, sai, ben, sis,
     pec, pecl, ib, cn, fd, pa, pend) in PROCS_BASE:
    PROCS.append(dict(Trilha=f"{mp} › {cod}", Macroprocesso=mp, Codigo=cod, Nome=nome,
        Descricao=desc, Objetivo=obj, Unidade_Organica_Responsavel=uor, Ponto_Focal_Nugep=pfn,
        Prioridade=pri, Complexidade=com, Maturidade=mat, Status_Mapeamento=sm, Percentual=None,
        Inicio_Mapeamento=im, Prazo_Previsto=pp, Data_Conclusao=dc_, Ultima_Atualizacao=ua,
        M1_Conhecer_Processo=m1, M2_Processo_Modelado=m2, M3_Subprocessos_Modelados=m3,
        M4_ASIS_Modelado=m4, M5_ASIS_Validado=m5, M6_Procedimento_Aprovado=m6,
        M7_Processo_Publicado=m7, M8_TOBE_Elaborado=m8, M9_TOBE_Aprovado=m9,
        M10_Processo_Transformado=m10, Fornecedores=forn, Entradas=ent, Saidas=sai,
        Beneficiarios=ben, Sistemas=sis, Processo_ECodevasf=pec, Processo_ECodevasf_Link=pecl,
        Imagem_Bizagi=ib, Competencias_Necessarias=cn, Fontes_Dados=fd, Proxima_Acao=pa, Pendencia=pend))

SUBS = []
for (mp, proc, cod, nome, vp, ordem, desc, obj, uor, reut, reut_em, ent, sai, sis, fd, ib) in SUBS_BASE:
    SUBS.append(dict(Trilha=f"{mp} › {proc} › {cod}", Macroprocesso=mp, Processo=proc, Codigo=cod,
        Nome=nome, Vinculo_Pai=vp, Ordem=ordem, Descricao=desc, Objetivo=obj,
        Unidade_Organica_Responsavel=uor, Reutilizavel=reut, Reutilizado_Em=reut_em,
        Entradas=ent, Saidas=sai, Sistemas=sis, Fontes_Dados=fd, Imagem_Bizagi=ib))

ATIVS = []
for (mp, proc, sub, cod, nome, vp, ordem, tipo, desc, ent, sai, sis) in ATIVS_BASE:
    trilha = f"{mp} › {proc} › " + (f"{sub} › " if sub else "") + cod
    ATIVS.append(dict(Trilha=trilha, Macroprocesso=mp, Processo=proc, Subprocesso=sub, Codigo=cod,
        Nome=nome, Vinculo_Pai=vp, Ordem=ordem, Tipo_Atividade=tipo, Descricao=desc,
        Entradas=ent, Saidas=sai, Sistemas=sis))

TAREFAS = []
for (mp, proc, sub, ativ, cod, nome, ordem, tipo, dur, obs) in TAREFAS_BASE:
    trilha = f"{mp} › {proc} › " + (f"{sub} › " if sub else "") + f"{ativ} › {cod}"
    TAREFAS.append(dict(Trilha=trilha, Macroprocesso=mp, Processo=proc, Subprocesso=sub, Atividade=ativ,
        Codigo=cod, Nome=nome, Ordem=ordem, Tipo_Tarefa=tipo, Duracao_Estimada=dur, Observacoes=obs))

DOCS = []
for (id_, tit, tipo, vc, ver, data, sit, ato, link, obs) in DOCS_BASE:
    DOCS.append(dict(ID=id_, Titulo=tit, Tipo_Documento=tipo, Vinculo_Nivel=_nivel(vc), Vinculo_Codigo=vc,
        Versao=ver, Data=data, Situacao=sit, Ato_Aprovacao=ato, Link=link, Observacoes=obs))

RISCOS = []
for (id_, vc, desc, cat, fat, cron, dtid, prob, imp, resp, ctrl, prazo, resp2, status) in RISCOS_BASE:
    p, i = float(prob or 0), float(imp or 0)
    nivel = int(p * i) if prob and imp else None
    classe = None
    if nivel is not None:
        classe = "Extremo" if nivel >= 20 else "Alto" if nivel >= 12 else "Moderado" if nivel >= 5 else "Baixo"
    RISCOS.append(dict(ID=id_, Vinculo_Nivel=_nivel(vc), Vinculo_Codigo=vc, Descricao_Risco=desc,
        Categoria=cat, Fatores=fat, Cronograma_Risco=cron, Data_Identificacao=dtid,
        Probabilidade_1a5=prob, Impacto_1a5=imp, Nivel_PxI=nivel, Classificacao=classe, Resposta=resp,
        Controles_Tratamento=ctrl, Prazo_Tratamento=prazo, Responsavel=resp2, Status=status))

METRICAS = []
for (id_, nome, vc, cat, formula, crit, un, pol, meta, per, fonte, obs) in METRICAS_BASE:
    METRICAS.append(dict(ID=id_, Nome=nome, Vinculo_Nivel=_nivel(vc), Vinculo_Codigo=vc, Categoria=cat,
        Descricao_Formula=formula, Criterios_Desempenho=crit, Unidade=un, Polaridade=pol, Meta=meta,
        Periodicidade=per, Fonte=fonte, Observacoes=obs))

PAPEIS = []
for (id_, vc, papel, env, up) in PAPEIS_BASE:
    PAPEIS.append(dict(ID=id_, Vinculo_Nivel=_nivel(vc), Vinculo_Codigo=vc, Papel=papel, Envolvimento=env, Unidade_Pessoa=up))

REGRAS = []
for (id_, nome, vc, tipo, desc, fonte) in REGRAS_BASE:
    REGRAS.append(dict(ID=id_, Nome=nome, Vinculo_Nivel=_nivel(vc), Vinculo_Codigo=vc, Tipo_Regra=tipo,
        Descricao=desc, Fonte_Normativa=fonte))


def _linha(d, aba):
    return [d.get(k) for k in ESQ.colunas(aba)]


MCOLS = ["M1_Conhecer_Processo", "M2_Processo_Modelado", "M3_Subprocessos_Modelados", "M4_ASIS_Modelado",
         "M5_ASIS_Validado", "M6_Procedimento_Aprovado", "M7_Processo_Publicado", "M8_TOBE_Elaborado",
         "M9_TOBE_Aprovado", "M10_Processo_Transformado"]


def _percentual(d):
    pontos = total = 0
    for k in MCOLS:
        v = d.get(k)
        if v == "Sim":
            pontos += 1; total += 1
        elif v == "Em andamento":
            pontos += 0.5; total += 1
        elif v == "Não":
            total += 1
    return round(pontos / total, 4) if total else 0


for d in PROCS:
    d["Percentual"] = _percentual(d)

# ──────────────────────────────── Workbook ──────────────────────────────
wb = Workbook()
ls = wb.active
ls.title = "Listas"

FK_TODOS_CODIGOS = ([p["Codigo"] for p in PROCS] + [s["Codigo"] for s in SUBS]
                     + [a["Codigo"] for a in ATIVS] + [t["Codigo"] for t in TAREFAS])
listas = dict(ESQ.LISTAS_OPCOES)

for j, (nome, itens) in enumerate(listas.items(), start=1):
    c = ls.cell(row=1, column=j, value=nome)
    c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = AL_HEAD; c.border = BORDA
    largura = min(max(len(nome), max((len(x) for x in itens), default=8)) + 3, 40)
    ls.column_dimensions[get_column_letter(j)].width = largura
    for i, item in enumerate(itens, start=2):
        cc = ls.cell(row=i, column=j, value=item)
        cc.font = F_CELL; cc.border = BORDA
ls.freeze_panes = "A2"
ls.sheet_properties.tabColor = COR_GUIA["Listas"]
ls.sheet_view.zoomScale = 90


def ref(nome):
    j = list(listas).index(nome) + 1
    coluna = get_column_letter(j)
    n = len(listas[nome])
    return f"=Listas!${coluna}$2:${coluna}${n + 1}"


REF_SIGLAS = f"=Siglas!$A$2:$A${len(SIGLAS) + 1}"
REF_NUGEP_NOME = f"=NUGEP!$B$2:$B${len(CONTEUDO.NUGEP) + 1}"

# ---- Macroprocessos ----
mp = wb.create_sheet("Macroprocessos")
cabecalho(mp, "Macroprocessos", formula_cols={ESQ.colunas("Macroprocessos").index("Trilha") + 1})
MACROS_D = [dict(Codigo=m[0], Nome=m[1], Categoria=m[2], Ordem=m[3], Unidade_Organica_Responsavel=m[4],
    Descricao=m[5], Objetivo=m[6], Entregas=m[7], Beneficiarios=m[8], Partes_Interessadas=m[9],
    Sistemas=m[10], Imagem_Bizagi=m[11], Observacoes=m[12], Trilha=m[0]) for m in MACROS]
escreve(mp, [_linha(d, "Macroprocessos") for d in MACROS_D], wrap_cols={2,6,7,8,9,10,11,12,13}, center_cols={1,3,4,5})
mp.freeze_panes = "B2"
lit_formula(mp, "Macroprocessos", "Trilha", MACROS_D)
dv(mp, col("Macroprocessos", "Categoria"), ref("Categoria_Macroprocesso"))
dv(mp, col("Macroprocessos", "Unidade_Organica_Responsavel"), REF_SIGLAS)

# ---- Processos ----
pr = wb.create_sheet("Processos")
cabecalho(pr, "Processos", formula_cols={ESQ.colunas("Processos").index("Trilha") + 1, ESQ.colunas("Processos").index("Percentual") + 1})
escreve(pr, [_linha(d, "Processos") for d in PROCS],
        wrap_cols={4,5,6,29,30,31,32,33,35,36,37,38,39,40},
        center_cols={2,3,9,10,11,12,13,14,15,16,17} | set(range(18,28)) | {34})
pr.freeze_panes = "D2"
lit_formula(pr, "Processos", "Trilha", PROCS)
_c1p, _c10p = col("Processos", "M1_Conhecer_Processo"), col("Processos", "M10_Processo_Transformado")
def _pct_formula(r):
    rng = f"{_c1p}{r}:{_c10p}{r}"
    partes = "+".join(f'IF({col("Processos", k)}{r}="Sim",1,IF({col("Processos", k)}{r}="Em andamento",0.5,0))' for k in MCOLS)
    denom = "+".join(f'IF(OR({col("Processos", k)}{r}="Sim",{col("Processos", k)}{r}="Não",{col("Processos", k)}{r}="Em andamento"),1,0)' for k in MCOLS)
    return f"=IF(({denom})=0,0,({partes})/({denom}))"
aplicar_formula(pr, "Processos", "Percentual", _pct_formula, number_format="0%")
dv(pr, col("Processos", "Macroprocesso"), f"=Macroprocessos!$A$2:$A${len(MACROS) + 1}")
dv(pr, col("Processos", "Unidade_Organica_Responsavel"), REF_SIGLAS)
dv(pr, col("Processos", "Ponto_Focal_Nugep"), REF_NUGEP_NOME)
dv(pr, col("Processos", "Prioridade"), ref("Prioridade"))
dv(pr, col("Processos", "Complexidade"), ref("Complexidade"))
dv(pr, col("Processos", "Maturidade"), ref("Maturidade_Processo"))
dv(pr, col("Processos", "Status_Mapeamento"), ref("Status_Mapeamento"))
for k in MCOLS:
    dv(pr, col("Processos", k), ref("Status_Marco"))

# ---- Subprocessos ----
sp = wb.create_sheet("Subprocessos")
cabecalho(sp, "Subprocessos", formula_cols={i for i in (ESQ.colunas("Subprocessos").index("Trilha") + 1, ESQ.colunas("Subprocessos").index("Macroprocesso") + 1, ESQ.colunas("Subprocessos").index("Processo") + 1)})
escreve(sp, [_linha(d, "Subprocessos") for d in SUBS], wrap_cols={5,8,9,12,13,14,15,16,17}, center_cols={1,2,3,4,6,7,10})
sp.freeze_panes = "E2"
lit_formula(sp, "Subprocessos", "Trilha", SUBS)
lit_formula(sp, "Subprocessos", "Macroprocesso", SUBS)
lit_formula(sp, "Subprocessos", "Processo", SUBS)
dv(sp, col("Subprocessos", "Unidade_Organica_Responsavel"), REF_SIGLAS)
dv(sp, col("Subprocessos", "Reutilizavel"), ref("Sim_Nao"))

# ---- Atividades ----
at = wb.create_sheet("Atividades")
cols_at_formula = {ESQ.colunas("Atividades").index(k) + 1 for k in ("Trilha", "Macroprocesso", "Processo", "Subprocesso")}
cabecalho(at, "Atividades", formula_cols=cols_at_formula)
escreve(at, [_linha(d, "Atividades") for d in ATIVS], wrap_cols={1,10,11,12,13}, center_cols={2,3,4,5,7,8})
at.freeze_panes = "F2"
lit_formula(at, "Atividades", "Trilha", ATIVS)
lit_formula(at, "Atividades", "Macroprocesso", ATIVS)
lit_formula(at, "Atividades", "Processo", ATIVS)
lit_formula(at, "Atividades", "Subprocesso", ATIVS)
dv(at, col("Atividades", "Tipo_Atividade"), ref("Tipo_Atividade"))

# ---- Tarefas ----
tf = wb.create_sheet("Tarefas")
cols_tf_formula = {ESQ.colunas("Tarefas").index(k) + 1 for k in ("Trilha", "Macroprocesso", "Processo", "Subprocesso")}
cabecalho(tf, "Tarefas", formula_cols=cols_tf_formula)
escreve(tf, [_linha(d, "Tarefas") for d in TAREFAS], wrap_cols={1,11}, center_cols={2,3,4,5,6,8,9,10})
tf.freeze_panes = "F2"
lit_formula(tf, "Tarefas", "Trilha", TAREFAS)
lit_formula(tf, "Tarefas", "Macroprocesso", TAREFAS)
lit_formula(tf, "Tarefas", "Processo", TAREFAS)
lit_formula(tf, "Tarefas", "Subprocesso", TAREFAS)
dv(tf, col("Tarefas", "Tipo_Tarefa"), ref("Tipo_Tarefa"))

# ---- Documentos ----
dc = wb.create_sheet("Documentos")
cabecalho(dc, "Documentos", formula_cols={ESQ.colunas("Documentos").index("Vinculo_Nivel") + 1})
escreve(dc, [_linha(d, "Documentos") for d in DOCS], wrap_cols={2,9,10,11}, center_cols={1,3,4,5,6,7,8})
dc.freeze_panes = "B2"
aplicar_formula(dc, "Documentos", "Vinculo_Nivel", lambda r: nivel_formula(f'{col("Documentos", "Vinculo_Codigo")}{r}'))
dv(dc, col("Documentos", "Tipo_Documento"), ref("Tipo_Documento"))
dv(dc, col("Documentos", "Situacao"), ref("Situacao_Documento"))

# ---- Riscos ----
rk = wb.create_sheet("Riscos")
cabecalho(rk, "Riscos", formula_cols={ESQ.colunas("Riscos").index(k) + 1 for k in ("Vinculo_Nivel", "Nivel_PxI", "Classificacao")})
escreve(rk, [_linha(d, "Riscos") for d in RISCOS], wrap_cols={4,6,7,14,16}, center_cols={1,2,3,5,8,9,10,11,12,13,15,17})
rk.freeze_panes = "C2"
aplicar_formula(rk, "Riscos", "Vinculo_Nivel", lambda r: nivel_formula(f'{col("Riscos", "Vinculo_Codigo")}{r}'))
_cp, _ci = col("Riscos", "Probabilidade_1a5"), col("Riscos", "Impacto_1a5")
aplicar_formula(rk, "Riscos", "Nivel_PxI", lambda r: f"={_cp}{r}*{_ci}{r}")
_cn = col("Riscos", "Nivel_PxI")
aplicar_formula(rk, "Riscos", "Classificacao", lambda r: f'=IF({_cn}{r}>=20,"Extremo",IF({_cn}{r}>=12,"Alto",IF({_cn}{r}>=5,"Moderado","Baixo")))')
dv(rk, col("Riscos", "Categoria"), ref("Categoria_Risco"))
dv(rk, col("Riscos", "Resposta"), ref("Resposta_Risco"))
dv(rk, col("Riscos", "Status"), ref("Status_Risco"))

# ---- Metricas ----
mt = wb.create_sheet("Metricas")
cabecalho(mt, "Metricas", formula_cols={ESQ.colunas("Metricas").index("Vinculo_Nivel") + 1})
escreve(mt, [_linha(d, "Metricas") for d in METRICAS], wrap_cols={2,6,7,12,13}, center_cols={1,3,4,5,8,9,10,11})
mt.freeze_panes = "B2"
aplicar_formula(mt, "Metricas", "Vinculo_Nivel", lambda r: nivel_formula(f'{col("Metricas", "Vinculo_Codigo")}{r}'))
dv(mt, col("Metricas", "Polaridade"), ref("Polaridade"))
dv(mt, col("Metricas", "Periodicidade"), ref("Periodicidade"))

# ---- Medicoes ----
md = wb.create_sheet("Medicoes")
cabecalho(md, "Medicoes")
escreve(md, MEDICOES, wrap_cols={5}, center_cols={1,2,3,4})
md.freeze_panes = "B2"
dv(md, col("Medicoes", "Metrica_ID"), f"=Metricas!$A$2:$A${len(METRICAS) + 1}")

# ---- Papeis ----
pp = wb.create_sheet("Papeis")
cabecalho(pp, "Papeis", formula_cols={ESQ.colunas("Papeis").index("Vinculo_Nivel") + 1})
escreve(pp, [_linha(d, "Papeis") for d in PAPEIS], wrap_cols={4,6}, center_cols={1,2,3,5})
pp.freeze_panes = "B2"
aplicar_formula(pp, "Papeis", "Vinculo_Nivel", lambda r: nivel_formula(f'{col("Papeis", "Vinculo_Codigo")}{r}'))
dv(pp, col("Papeis", "Envolvimento"), ref("Envolvimento_RACI"))

# ---- Regras ----
rg = wb.create_sheet("Regras")
cabecalho(rg, "Regras", formula_cols={ESQ.colunas("Regras").index("Vinculo_Nivel") + 1})
escreve(rg, [_linha(d, "Regras") for d in REGRAS], wrap_cols={2,6,7}, center_cols={1,3,4,5})
rg.freeze_panes = "B2"
aplicar_formula(rg, "Regras", "Vinculo_Nivel", lambda r: nivel_formula(f'{col("Regras", "Vinculo_Codigo")}{r}'))
dv(rg, col("Regras", "Tipo_Regra"), ref("Tipo_Regra"))

# ---- Cultura_Processos / Iniciativas ----
cu = wb.create_sheet("Cultura_Processos")
cabecalho(cu, "Cultura_Processos")
escreve(cu, CULTURA, wrap_cols={3,5}, center_cols={1,2,4})
cu.freeze_panes = "B2"
dv(cu, col("Cultura_Processos", "Situacao"), ref("Situacao_Cultura"))

ini = wb.create_sheet("Iniciativas")
cabecalho(ini, "Iniciativas")
escreve(ini, INICIATIVAS, wrap_cols={3,5,6}, center_cols={1,2,4})
ini.freeze_panes = "B2"
dv(ini, col("Iniciativas", "Tipo"), ref("Tipo_Iniciativa"))

# ---- Competencias / Jornada / Repositorio / NUGEP / Glossario / FAQ / Siglas / Parametros ----
ct = wb.create_sheet("Competencias")
cabecalho(ct, "Competencias")
escreve(ct, CONTEUDO.COMPETENCIAS, wrap_cols={2,4}, center_cols={1,3})
ct.freeze_panes = "B2"

jn = wb.create_sheet("Jornada")
cabecalho(jn, "Jornada")
escreve(jn, CONTEUDO.JORNADA, wrap_cols={3,5,6,7,8,9}, center_cols={1,2,4})
jn.freeze_panes = "C2"

rp = wb.create_sheet("Repositorio")
cabecalho(rp, "Repositorio")
escreve(rp, CONTEUDO.REPOSITORIO, wrap_cols={4,5,6,7}, center_cols={1,2,3,8})
rp.freeze_panes = "B2"
dv(rp, col("Repositorio", "Categoria"), ref("Categoria_Repositorio"))
dv(rp, col("Repositorio", "Fase_Ciclo"), ref("Fase_Instrumento"), strict=False)

ng = wb.create_sheet("NUGEP")
cabecalho(ng, "NUGEP", formula_cols={ESQ.colunas("NUGEP").index("Unidade_Nome") + 1})
escreve(ng, CONTEUDO.NUGEP, wrap_cols={2,3,5,6,8}, center_cols={1,4,7,9})
ng.freeze_panes = "B2"
_csig = col("NUGEP", "Unidade_Sigla")
aplicar_formula(ng, "NUGEP", "Unidade_Nome", lambda r: f'=IFERROR(VLOOKUP({_csig}{r},Siglas!$A:$B,2,0),"")')
dv(ng, col("NUGEP", "Unidade_Sigla"), REF_SIGLAS)

gl = wb.create_sheet("Glossario")
cabecalho(gl, "Glossario")
escreve(gl, CONTEUDO.GLOSSARIO, wrap_cols={1,2,3})
gl.freeze_panes = "A2"

fq = wb.create_sheet("FAQ")
cabecalho(fq, "FAQ")
escreve(fq, CONTEUDO.FAQ, wrap_cols={3,4}, center_cols={1,2})
fq.freeze_panes = "C2"
dv(fq, col("FAQ", "Categoria"), ref("Categoria_FAQ"))

sg = wb.create_sheet("Siglas")
cabecalho(sg, "Siglas")
escreve(sg, SIGLAS, wrap_cols={2}, center_cols={1})
sg.freeze_panes = "A2"

pm = wb.create_sheet("Parametros")
cabecalho(pm, "Parametros")
escreve(pm, CONTEUDO.PARAMETROS, wrap_cols={2}, center_cols={1})
pm.freeze_panes = "A2"

# ──────────────────────────────── LEIA-ME ───────────────────────────────
lm = wb.create_sheet("LEIA-ME")
lm.sheet_properties.tabColor = COR_GUIA["LEIA-ME"]
lm.sheet_view.showGridLines = False
lm.column_dimensions["A"].width = 3
lm.column_dimensions["B"].width = 30
lm.column_dimensions["C"].width = 100


def titulo(r, txt, size=14, cor=AZUL_ESCURO):
    c = lm.cell(row=r, column=2, value=txt)
    c.font = Font(name=FONTE, size=size, bold=True, color=cor)


def linha(r, rotulo, texto):
    a = lm.cell(row=r, column=2, value=rotulo)
    a.font = Font(name=FONTE, size=10, bold=True)
    a.alignment = AL_TOP
    b = lm.cell(row=r, column=3, value=texto)
    b.font = F_CELL
    b.alignment = AL_WRAP


titulo(2, "Painel de Processos — Base de Dados (DADOS FICTÍCIOS)", 16)
lm.cell(row=3, column=2, value="Fonte única de dados do painel.").font = Font(name=FONTE, size=10, italic=True)
linha(5, "Códigos reiniciam por vínculo", "PP-/SP-/AT-/TR- recomeçam em 01 a cada novo Macroprocesso/Processo/"
      "Subprocesso/Atividade — não são mais únicos sozinhos. A coluna Trilha (1ª de cada aba da hierarquia) "
      "é o identificador de leitura: mostra o caminho completo, ex. \"MS-01 › PP-01 › SP-03 › SP-01\". "
      "Reutilizado_Em e qualquer referência entre abas usam o caminho completo (ex. MG-01/PP-03/SP-03).")
linha(6, "Trilha e rollups", "Calculadas no momento da geração — não são VLOOKUP vivo por código (o código "
      "deixou de ser único). Reordenar ou trocar o vínculo de um item exige rodar o gerador de novo.")
linha(7, "Marco \"Em andamento\"", "M1–M10 aceitam Sim / Em andamento / Não / Não se aplica. Em andamento "
      "vale meia nota no Percentual.")
linha(8, "Datas", "Formato dd/mm/aaaa.")

ordem_final = ["LEIA-ME", "Macroprocessos", "Processos", "Subprocessos", "Atividades", "Tarefas",
               "Documentos", "Riscos", "Metricas", "Medicoes", "Papeis", "Regras",
               "Cultura_Processos", "Iniciativas", "Competencias", "Jornada", "Repositorio",
               "NUGEP", "Glossario", "FAQ", "Siglas", "Parametros", "Listas"]
wb._sheets = [wb[n] for n in ordem_final]
wb.active = 0

from pathlib import Path

DESTINO = Path(__file__).resolve().parent.parent / "data" / "painel-processos-dados.xlsx"
wb.save(DESTINO)

from cachear_formulas import cachear  # noqa: E402

print(f"Planilha gerada com sucesso ({cachear(DESTINO)} fórmulas com valor calculado).")
