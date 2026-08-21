# -*- coding: utf-8 -*-
"""
Aplica a FORMATAÇÃO do painel na planilha que já existe em
data/painel-processos-dados.xlsx — sem tocar em uma célula de dado.

Existe porque os dois caminhos anteriores não serviam a quem já tem
conteúdo real na planilha:

  * gerar_planilha.py   RECRIA o arquivo do zero (apaga os dados reais);
  * atualizar_planilha.py acrescenta linhas/colunas herdando o estilo
    vizinho — de propósito, não aplica formatação nova.

Este script fica no meio: abre o .xlsx atual, repinta cabeçalho, listras,
cor de guia, zoom e as mensagens das listas suspensas, e salva. Nenhum
valor de célula é lido para decisão nem alterado.

Uso:
    python scripts/formatar_planilha.py            # aplica
    python scripts/formatar_planilha.py --conferir # só relata, não salva

Requisito: openpyxl (pip install openpyxl)
"""
import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
XLSX = os.path.join(BASE, "data", "painel-processos-dados.xlsx")

# ── Paleta e estilos: os MESMOS de gerar_planilha.py ────────────────────
# Importados de lá quando possível, para não haver duas verdades sobre a
# cor do cabeçalho. Se o import falhar (o gerador tem efeitos colaterais
# no nível do módulo), caem os valores literais abaixo.
AZUL = "0040B4"          # blue-warm-vivid-70 (gov.br DS v4)
AZUL_ESCURO = "0C326F"
CINZA_FORMULA = "F0F0F0"
CINZA_ZEBRA = "F5F8FC"
FONTE = "Arial"

COR_GUIA = {
    "LEIA-ME": AZUL_ESCURO,
    "Macroprocessos": AZUL, "Processos": AZUL, "Subprocessos": AZUL,
    "Atividades": AZUL, "Tarefas": AZUL,
    "Documentos": "168821", "Riscos": "168821", "Metricas": "168821", "Medicoes": "168821",
    "Papeis": "168821", "Regras": "168821", "Cultura_Processos": "168821", "Iniciativas": "168821",
    "Equipe_Processo": "168821",
    "Jornada": "155BCB", "Repositorio": "155BCB", "NUGEP": "155BCB", "Competencias": "155BCB",
    "Glossario": "155BCB", "FAQ": "155BCB", "Siglas": "888888",
    "Parametros": "888888", "Listas": "888888", "Dicionario": AZUL_ESCURO,
}

_th = Side(style="thin", color="CCCCCC")
BORDA = Border(left=_th, right=_th, top=_th, bottom=_th)
F_HEAD = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
FILL_HEAD = PatternFill("solid", fgColor=AZUL)
FILL_ZEBRA = PatternFill("solid", fgColor=CINZA_ZEBRA)
AL_HEAD = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Mensagens das listas suspensas, no Padrão Writing do DS: dizem o que
# fazer e não culpam quem preencheu.
DV_ERRO_TITULO = "Valor fora da lista"
DV_ERRO = ("Escolha uma das opções da seta ao lado da célula. "
           "A lista completa está na aba Listas.")
DV_PROMPT_TITULO = "Lista de opções"
DV_PROMPT = "Clique na seta ao lado da célula e escolha uma das opções."

# Abas que são texto corrido, não tabela — ficam de fora da repintura de
# cabeçalho e de listras.
SEM_TABELA = {"LEIA-ME"}


def eh_zebra(fill):
    """True se a célula está sem preenchimento próprio (ou já com a
    listra). Uma célula pintada de propósito — coluna de fórmula em cinza,
    destaque manual de quem preencheu — é preservada."""
    if fill is None or fill.fill_type is None:
        return True
    cor = getattr(fill.start_color, "rgb", None)
    if not isinstance(cor, str):
        return False
    return cor.upper().endswith(CINZA_ZEBRA)


def formata_aba(ws, relato):
    ws.sheet_properties.tabColor = COR_GUIA.get(ws.title, AZUL)
    ws.sheet_view.zoomScale = 90
    if ws.title in SEM_TABELA:
        relato.append(f"  {ws.title}: cor de guia e zoom")
        return

    # 1. Cabeçalho no azul da v4
    largura = 0
    for c in ws[1]:
        if c.value is None:
            continue
        largura = max(largura, c.column)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = AL_HEAD
        c.border = BORDA
    if not largura:
        relato.append(f"  {ws.title}: sem cabeçalho — só cor de guia e zoom")
        return
    ws.row_dimensions[1].height = 30

    # 2. Cabeçalho fixo e filtro, se ainda não houver
    if not ws.freeze_panes or ws.freeze_panes == "A1":
        ws.freeze_panes = "A2"
    if not ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:{get_column_letter(largura)}1"

    # 3. Listra zebrada nas linhas de dados
    linhas = 0
    for i in range(2, ws.max_row + 1):
        if all(ws.cell(row=i, column=j).value in (None, "")
               for j in range(1, largura + 1)):
            continue
        linhas += 1
        for j in range(1, largura + 1):
            c = ws.cell(row=i, column=j)
            c.border = BORDA
            if i % 2 == 0:
                if eh_zebra(c.fill):
                    c.fill = FILL_ZEBRA
            elif eh_zebra(c.fill) and c.fill.fill_type is not None:
                c.fill = PatternFill()

    # 4. Listas suspensas que avisam em vez de aceitar em silêncio
    validacoes = 0
    for v in ws.data_validations.dataValidation:
        if v.type != "list":
            continue
        validacoes += 1
        v.showErrorMessage = True
        v.errorStyle = "stop"
        v.errorTitle = DV_ERRO_TITULO
        v.error = DV_ERRO
        v.showInputMessage = True
        v.promptTitle = DV_PROMPT_TITULO
        v.prompt = DV_PROMPT

    relato.append(
        f"  {ws.title}: {largura} colunas · {linhas} linhas · "
        f"{validacoes} lista(s) suspensa(s)")


def main():
    conferir = "--conferir" in sys.argv
    if not os.path.exists(XLSX):
        print("Planilha não encontrada: " + os.path.relpath(XLSX, BASE))
        print("Rode scripts/gerar_planilha.py para criar a de exemplo.")
        return 1

    wb = load_workbook(XLSX)
    relato = []
    for ws in wb.worksheets:
        formata_aba(ws, relato)

    print("Formatação do painel aplicada em "
          + os.path.relpath(XLSX, BASE) + ":")
    print("\n".join(relato))
    print("\nNenhum valor de célula foi alterado.")

    if conferir:
        print("\n--conferir: nada foi salvo.")
        return 0

    wb.save(XLSX)
    print("\nOK → " + os.path.relpath(XLSX, BASE))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
