# -*- coding: utf-8 -*-
"""
Atualiza data/painel-processos-dados.xlsx NO LUGAR, sem recriar a planilha:
soma colunas/abas novas preservando dados já preenchidos, formatação,
fórmulas e validações existentes.

Rode com:
    python scripts/atualizar_planilha.py

Idempotente: cada mudança confere se já foi aplicada antes de agir, para
rodar de novo não duplicar nada. Escreva aqui o PRÓXIMO incremento quando
uma auditoria de metodologia pedir uma coluna ou aba nova — não use
gerar_planilha.py para a planilha em uso (ele recria do zero, com dados
fictícios, a partir de scripts/esquema_planilha.py).

Nenhuma migração pendente no momento — todas as auditorias até aqui já
foram incorporadas à planilha em uso (ver pendencias.md). Os utilitários
abaixo ficam prontos para o próximo incremento.
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CAMINHO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                       "data", "painel-processos-dados.xlsx")

# Mesmas constantes visuais de gerar_planilha.py — mantenha as duas em sincronia.
AZUL = "0040B4"
CINZA_ZEBRA = "F5F8FC"
FONTE = "Arial"
th = Side(style="thin", color="CCCCCC")
BORDA = Border(left=th, right=th, top=th, bottom=th)
F_HEAD = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
F_CELL = Font(name=FONTE, size=10)
FILL_HEAD = PatternFill("solid", fgColor=AZUL)
FILL_ZEBRA = PatternFill("solid", fgColor=CINZA_ZEBRA)
AL_HEAD = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_WRAP = Alignment(vertical="top", wrap_text=True)
AL_TOP = Alignment(vertical="top")
AL_CENTER = Alignment(horizontal="center", vertical="top")


def cabecalho(ws, headers, widths, tab_color=AZUL):
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = AL_HEAD; c.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.zoomScale = 90


def escreve(ws, linhas, wrap_cols=(), center_cols=()):
    for i, linha in enumerate(linhas, start=2):
        for j, v in enumerate(linha, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = F_CELL; c.border = BORDA
            if i % 2 == 0:
                c.fill = FILL_ZEBRA
            c.alignment = AL_WRAP if j in wrap_cols else (AL_CENTER if j in center_cols else AL_TOP)


def col_by_header(ws, nome, linha=1):
    """Índice (1-based) da coluna cujo cabeçalho é `nome`, ou None."""
    for cell in ws[linha]:
        if cell.value == nome:
            return cell.column
    return None


def listas_ref(wb, nome_lista):
    """Referência de validação para uma coluna nomeada da aba Listas, ou None
    se a aba Listas não tiver essa coluna (planilha mais antiga)."""
    if "Listas" not in wb.sheetnames:
        return None
    ls = wb["Listas"]
    col = col_by_header(ls, nome_lista)
    if not col:
        return None
    n = 1
    while ls.cell(row=n + 1, column=col).value not in (None, ""):
        n += 1
    letra = get_column_letter(col)
    return f"=Listas!${letra}$2:${letra}${n + 1}"


def siglas_ref(wb):
    """Referência de validação para a aba Siglas (coluna A, até a última
    linha preenchida) — mesmo padrão de listas_ref(), mas para a aba Siglas,
    que os campos de unidade orgânica referenciam diretamente (não vai pela
    aba Listas)."""
    if "Siglas" not in wb.sheetnames:
        return None
    ws = wb["Siglas"]
    n = 1
    while ws.cell(row=n + 1, column=1).value not in (None, ""):
        n += 1
    return f"=Siglas!$A$2:$A${n + 1}"


def dv(ws, col_letter, ref, ultima_linha=300, strict=True):
    if not ref:
        return
    v = DataValidation(type="list", formula1=ref, allow_blank=True,
                        showErrorMessage=True, errorStyle="stop" if strict else "warning",
                        errorTitle="Valor fora da lista",
                        error="Escolha uma das opções da lista. A lista completa está na aba Listas.",
                        showInputMessage=True, promptTitle="Lista de opções",
                        prompt="Clique na seta ao lado da célula e escolha uma das opções.")
    ws.add_data_validation(v)
    v.add(f"{col_letter}2:{col_letter}{ultima_linha}")


def principal():
    wb = load_workbook(CAMINHO)
    mudou = []

    # Gestor(a) do processo (nome/e-mail/telefone/unidade orgânica) e aba
    # Equipe_Gerenciamento_Processos — pedido do usuário (ago/2026). Colunas
    # novas vão ao final da aba Processos (não entre as existentes), para
    # não deslocar letra de coluna de nada já preenchido.
    ws_proc = wb["Processos"]
    if not col_by_header(ws_proc, "Gestor_Nome"):
        ultima_col = ws_proc.max_column
        novas = [("Gestor_Nome", 24), ("Gestor_Email", 30),
                 ("Gestor_Telefone", 16), ("Gestor_Unidade_Organica", 16)]
        for k, (nome, largura) in enumerate(novas, start=1):
            j = ultima_col + k
            c = ws_proc.cell(row=1, column=j, value=nome)
            c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = AL_HEAD; c.border = BORDA
            ws_proc.column_dimensions[get_column_letter(j)].width = largura
            for i in range(2, ws_proc.max_row + 1):
                cc = ws_proc.cell(row=i, column=j)
                cc.border = BORDA
                cc.alignment = AL_WRAP if nome == "Gestor_Nome" else AL_TOP
                if i % 2 == 0:
                    cc.fill = FILL_ZEBRA
        ref = siglas_ref(wb)
        if ref:
            dv(ws_proc, get_column_letter(ultima_col + 4), ref)
        mudou.append("Processos: colunas Gestor_Nome, Gestor_Email, Gestor_Telefone, Gestor_Unidade_Organica")

    if "Equipe_Gerenciamento_Processos" not in wb.sheetnames:
        ws_eq = wb.create_sheet("Equipe_Gerenciamento_Processos")
        cabecalho(ws_eq, ["Ordem", "Nome", "Email", "Telefone", "Unidade_Organica"],
                  [7, 28, 32, 16, 16])
        ref = siglas_ref(wb)
        if ref:
            dv(ws_eq, "E", ref)
        mudou.append("Nova aba Equipe_Gerenciamento_Processos (Ordem, Nome, Email, Telefone, Unidade_Organica)")

    # Trilha com "›" em vez de "/" nos vínculos (padronização pedida pelo
    # usuário) — Vinculo_Codigo em Documentos/Riscos/Metricas/Papeis/Regras,
    # Reutilizado_Em em Subprocessos, Processos_Relacionados em Iniciativas.
    # Não toca sigla de unidade orgânica (AE/GPE/UNP) — esses campos não têm
    # "/" como separador de trilha, então não entram nesta lista de colunas.
    def normaliza_trilhas(valor):
        if not isinstance(valor, str) or "/" not in valor:
            return valor
        return "; ".join(
            " › ".join(seg.strip() for seg in parte.split("/") if seg.strip())
            for parte in valor.split(";")
        )

    campos_trilha = [
        ("Documentos", "Vinculo_Codigo"), ("Riscos", "Vinculo_Codigo"),
        ("Metricas", "Vinculo_Codigo"), ("Papeis", "Vinculo_Codigo"),
        ("Regras", "Vinculo_Codigo"), ("Subprocessos", "Reutilizado_Em"),
        ("Iniciativas", "Processos_Relacionados"),
    ]
    trilhas_corrigidas = 0
    for aba, coluna in campos_trilha:
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        col = col_by_header(ws, coluna)
        if not col:
            continue
        for i in range(2, ws.max_row + 1):
            cel = ws.cell(row=i, column=col)
            novo = normaliza_trilhas(cel.value)
            if novo != cel.value:
                cel.value = novo
                trilhas_corrigidas += 1
    if trilhas_corrigidas:
        mudou.append(f"Trilha com › em vez de / normalizada em {trilhas_corrigidas} célula(s) de vínculo")

    if mudou:
        wb.save(CAMINHO)
        print("Planilha atualizada:")
        for m in mudou:
            print(" -", m)
        # wb.save() aqui relê e regrava o arquivo com openpyxl, que nunca
        # escreve o valor calculado de uma fórmula (só a fórmula em si) —
        # mesmo em fórmulas que este script não tocou. Sem recachear,
        # Processos.Percentual, Documentos/Riscos/Metricas/Papeis/Regras
        # .Vinculo_Nivel, Riscos.Nivel_PxI/Classificacao e NUGEP.Unidade_Nome
        # voltam a sair vazios para qualquer leitor que não seja o Excel
        # (SheetJS incluso).
        print("\nRode em seguida, NESTA ORDEM (a primeira é obrigatória):")
        print("  python scripts/cachear_formulas.py")
        print("  python scripts/planilha_para_js.py")
    else:
        print("Nada para atualizar — a planilha já está em dia.")


if __name__ == "__main__":
    principal()
